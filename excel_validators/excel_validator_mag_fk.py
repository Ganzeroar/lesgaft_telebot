from openpyxl import load_workbook
import glob
import configurations
import os
import re

from file_not_valid_exception import File_not_valid
from excel_validator import Excel_validator


class Excel_validator_mag_fk(Excel_validator):

    def run_validator(self, route):
        work_files = glob.glob(str(route) + '/*.xlsx')
        for work_file_name in work_files:
            try:
                self.check_file_name(work_file_name)
                work_book = load_workbook(work_file_name)
                self.check_worksheet_names(work_book.sheetnames)
                self.check_structure(work_book, work_file_name)
                self.check_cells_with_lessons(work_book, work_file_name)
            finally:
                path = os.path.join(os.path.abspath(
                    os.path.dirname(__file__)), work_file_name)
                os.remove(path)
        return f'{work_file_name[87:]} валиден'

    def run_validator_for_excel_parser(self, route):
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        for work_file_name in work_files:
            self.check_file_name(work_file_name)
            work_book = load_workbook(work_file_name)
            self.check_worksheet_names(work_book.sheetnames)
            self.check_structure(work_book, work_file_name)
            self.check_cells_with_lessons(work_book, work_file_name)
            print(f'{work_file_name[87:]} валиден')

    def check_cells_with_lessons(self, work_book, work_file_name):
        constants = self.return_current_file_constants(work_file_name)
        number_of_groups = constants['number_of_groups']
        group_cell_constants = self.get_group_cell_constants(constants)

        for worksheet_name in work_book.sheetnames:
            if self.is_reason_to_skip(worksheet_name) == True:
                continue
            worksheet = work_book[worksheet_name]
            for number in range(number_of_groups):
                first_row = group_cell_constants[number][0]
                last_row = group_cell_constants[number][1]
                first_column = group_cell_constants[number][2]
                self.check_lesson_cells_using_constants(
                    worksheet, worksheet_name, first_row, last_row, first_column)

    def check_lesson_cells_using_constants(self, worksheet, worksheet_name, first_lesson_cell_row, last_lesson_cell_row, first_lesson_cell_column):
        for row in range(first_lesson_cell_row, last_lesson_cell_row):
            viewed_lesson_cell = worksheet.cell(
                row=row, column=first_lesson_cell_column)
            viewed_location_cell = worksheet.cell(
                row=row, column=first_lesson_cell_column + 1)
            viewed_teacher_cell = worksheet.cell(
                row=row, column=first_lesson_cell_column + 2)
            self.check_is_lesson_cell_correct(
                worksheet, worksheet_name, viewed_lesson_cell)
            self.check_is_location_cell_correct(
                worksheet, worksheet_name, viewed_location_cell)
            self.check_is_teacher_cell_correct(
                worksheet, worksheet_name, viewed_teacher_cell)

    def check_is_teacher_cell_correct(self, worksheet, worksheet_name, viewed_teacher_cell):
        if self.is_merged(worksheet, viewed_teacher_cell) == True:
            viewed_teacher_cell_value = self.get_merged_cell_value(
                worksheet, viewed_teacher_cell)
        else:
            viewed_teacher_cell_value = viewed_teacher_cell.value
        if viewed_teacher_cell_value == None:
            return
        if 'практика' in viewed_teacher_cell_value:
            return
        teachers = viewed_teacher_cell_value.split('\n')
        if len(teachers) == 1:
            if teachers[0] not in configurations.existing_teachers:
                raise File_not_valid(
                    f'Ошибка в ячейке в {viewed_teacher_cell.coordinate} в листе {worksheet_name} в преподавателе {teachers[0]}')
        if len(teachers) == 2:
            if teachers[0] not in configurations.existing_teachers:
                raise File_not_valid(
                    f'Ошибка в ячейке в {viewed_teacher_cell.coordinate} в листе {worksheet_name} в преподавателе {teachers[0]}')
            if teachers[1] not in configurations.existing_teachers:
                raise File_not_valid(
                    f'Ошибка в ячейке в {viewed_teacher_cell.coordinate} в листе {worksheet_name} в преподавателе {teachers[1]}')

    def check_is_location_cell_correct(self, worksheet, worksheet_name, viewed_lesson_cell):
        if self.is_merged(worksheet, viewed_lesson_cell) == True:
            viewed_location_cell_value = self.get_merged_cell_value(
                worksheet, viewed_lesson_cell)
        else:
            viewed_location_cell_value = viewed_lesson_cell.value
        if viewed_location_cell_value == None:
            return
        location = viewed_location_cell_value
        if 'практика' in location:
            return
        if location not in configurations.existing_locations:
            raise File_not_valid(
                f'Ошибка в ячейке в {viewed_lesson_cell.coordinate} в листе {worksheet_name} в локации "{location}"')

    def check_is_lesson_cell_correct(self, worksheet, worksheet_name, viewed_lesson_cell):
        if self.is_merged(worksheet, viewed_lesson_cell) == True:
            viewed_lesson_cell_value = self.get_merged_cell_value(
                worksheet, viewed_lesson_cell)
        else:
            viewed_lesson_cell_value = viewed_lesson_cell.value
        if viewed_lesson_cell_value == None:
            return
        lesson_and_lesson_type = viewed_lesson_cell_value.split('\n')
        if len(lesson_and_lesson_type) != 2:
            raise File_not_valid(
                f'Ошибка в ячейке в {viewed_lesson_cell.coordinate} в листе {worksheet_name} в предмете "{lesson_and_lesson_type}"')

        lesson = lesson_and_lesson_type[0]
        lesson_type = lesson_and_lesson_type[1] #в случае практики - даты
        if 'практика' in lesson:
            self.check_practice_cell(viewed_lesson_cell, worksheet_name, lesson, lesson_type)
            return
            
        if lesson not in configurations.existing_subjects:
            raise File_not_valid(
                f'Ошибка в ячейке в {viewed_lesson_cell.coordinate} в листе {worksheet_name} в предмете "{lesson_and_lesson_type}"')
        if lesson_type not in configurations.existing_type_of_subjects:
            raise File_not_valid(
                f'Ошибка в ячейке в {viewed_lesson_cell.coordinate} в листе {worksheet_name} в типе предмет "{lesson_type}"')

    def get_group_cell_constants(self, constants):
        number_of_groups = constants['number_of_groups']

        group_cell_constants = [
            [
                constants['first_group_first_lesson_cell_row'],
                constants['first_group_last_lesson_cell_row'],
                constants['first_group_first_lesson_cell_column'],
            ],
            [
                constants['second_group_first_lesson_cell_row'],
                constants['second_group_last_lesson_cell_row'],
                constants['second_group_first_lesson_cell_column'],
            ],
            [
                constants['third_group_first_lesson_cell_row'],
                constants['third_group_last_lesson_cell_row'],
                constants['third_group_first_lesson_cell_column'],
            ],
        ]
        if number_of_groups == 4:
            fourth_group_first_lesson_cell_row = constants['fourth_group_first_lesson_cell_row']
            fourth_group_last_lesson_cell_row = constants['fourth_group_last_lesson_cell_row']
            fourth_group_first_lesson_cell_column = constants['fourth_group_first_lesson_cell_column']
            group_cell_constants.append(
                [fourth_group_first_lesson_cell_row, fourth_group_last_lesson_cell_row, fourth_group_first_lesson_cell_column])

        return group_cell_constants

    def check_structure(self, work_book, work_file_name):
        for worksheet_name in work_book.sheetnames:
            if self.is_reason_to_skip(worksheet_name) == True:
                continue
            worksheet = work_book[worksheet_name]
            self.check_group_struct(worksheet, work_file_name, worksheet_name)
            self.check_date_struct(worksheet, work_file_name, worksheet_name)
            self.check_day_struct(worksheet, work_file_name, worksheet_name)
            self.check_time_struct(worksheet, work_file_name, worksheet_name)
        return 'Структура ОК\n'

    def check_group_struct(self, worksheet, work_file_name, worksheet_name):
        constants = self.return_current_file_constants(work_file_name)
        const_group_row = constants['group_row']
        number_of_groups = constants['number_of_groups']
        const_first_group_first_column = constants['first_group_first_column']
        const_first_group_last_column = constants['first_group_last_column']
        const_first_group_number = constants['first_group_number']
        const_second_group_first_column = constants['second_group_first_column']
        const_second_group_last_column = constants['second_group_last_column']
        const_second_group_number = constants['second_group_number']
        const_third_group_first_column = constants['third_group_first_column']
        const_third_group_last_column = constants['third_group_last_column']
        const_third_group_number = constants['third_group_number']

        for column in range(const_first_group_first_column, const_first_group_last_column + 1):
            viewed_group_cell = worksheet.cell(
                row=const_group_row, column=column)
            if self.is_merged(worksheet, viewed_group_cell):
                viewed_group_value = self.get_merged_cell_value(
                    worksheet, viewed_group_cell)
                if viewed_group_value != const_first_group_number:
                    raise File_not_valid(
                        f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')
            else:
                raise File_not_valid(
                    f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')

        for column in range(const_second_group_first_column, const_second_group_last_column + 1):
            viewed_group_cell = worksheet.cell(
                row=const_group_row, column=column)
            if self.is_merged(worksheet, viewed_group_cell):
                viewed_group_value = self.get_merged_cell_value(
                    worksheet, viewed_group_cell)
                if viewed_group_value != const_second_group_number:
                    raise File_not_valid(
                        f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')
            else:
                raise File_not_valid(
                    f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')

        for column in range(const_third_group_first_column, const_third_group_last_column + 1):
            viewed_group_cell = worksheet.cell(
                row=const_group_row, column=column)
            if self.is_merged(worksheet, viewed_group_cell):
                viewed_group_value = self.get_merged_cell_value(
                    worksheet, viewed_group_cell)
                if viewed_group_value != const_third_group_number:
                    raise File_not_valid(
                        f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')
            else:
                raise File_not_valid(
                    f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')

        if number_of_groups == 4:
            const_fourth_group_first_column = constants['fourth_group_first_column']
            const_fourth_group_last_column = constants['fourth_group_last_column']
            const_fourth_group_number = constants['fourth_group_number']
            for column in range(const_fourth_group_first_column, const_fourth_group_last_column + 1):
                viewed_group_cell = worksheet.cell(
                    row=const_group_row, column=column)
                if self.is_merged(worksheet, viewed_group_cell):
                    viewed_group_value = self.get_merged_cell_value(
                        worksheet, viewed_group_cell)
                    if viewed_group_value != const_fourth_group_number:
                        raise File_not_valid(
                            f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')
                else:
                    raise File_not_valid(
                        f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')

    def find_clear_file_name(self, file_name):
        if '1_mag_fk' in file_name:
            return 'mag_fk_1'
        elif '2_mag_fk' in file_name:
            return 'mag_fk_2'
        else:
            return None

    def return_current_file_constants(self, work_file_name):
        clear_file_name = self.find_clear_file_name(work_file_name)
        constants = configurations.timetable_constants[clear_file_name]
        return constants

    # TODO потенциально можно вынести в константы и объединить с другими
    def check_file_name(self, work_file_name):
        if '1_mag_fk' in work_file_name:
            return 'Имя файла ОК\n'
        elif '2_mag_fk' in work_file_name:
            return 'Имя файла ОК\n'
        else:
            raise File_not_valid(f'Ошибка в имени файла {work_file_name}')
