from openpyxl import load_workbook
import glob
import re
import configurations
import os

from file_not_valid_exception import File_not_valid
from excel_validators.excel_validator import Excel_validator


class Excel_validator_lovs_zovs(Excel_validator):

    def run_validator(self, route):
        work_files = glob.glob(str(route) + '/*.xlsx')
        for work_file_name in work_files:
            try:
                self.check_file_name(work_file_name)
                work_book = load_workbook(work_file_name)
                self.check_worksheet_names(work_book.sheetnames)
                self.check_structure(work_book, work_file_name)

                print(f'{work_file_name} валиден')
            finally:
                path = os.path.join(os.path.abspath(
                    os.path.dirname(__file__)), work_file_name)
                os.remove(path)
        return f'{work_file_name[87:]} валиден'

    def run_validator_for_excel_parser(self, route):
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        for work_file_name in work_files:
            try:
                self.check_file_name(work_file_name)
                work_book = load_workbook(work_file_name)
                self.check_worksheet_names(work_book.sheetnames)
                self.check_structure(work_book, work_file_name)
            except Exception as exception:
                raise File_not_valid(
                    f'{exception} в файле {work_file_name[36:]}')
            print(f'{work_file_name} валиден')

    def check_structure(self, work_book, work_file_name):
        constants = self.return_current_file_constants(work_file_name)

        for worksheet_name in work_book.sheetnames:
            if self.is_reason_to_skip(worksheet_name) == True:
                continue
            worksheet = work_book[worksheet_name]
            self.check_group_struct(worksheet, work_file_name, worksheet_name)
            self.check_time_struct(worksheet, worksheet_name, constants)
            self.check_date_struct(worksheet, worksheet_name, constants)
            self.check_day_struct(worksheet, worksheet_name, constants)
        return 'Структура ОК\n'

    def check_group_struct(self, worksheet, work_file_name, worksheet_name):
        constants = self.return_current_file_constants(work_file_name)
        const_group_row = constants['group_row']
        const_first_group_column = constants['first_group_column']
        const_last_group_column = constants['last_group_column']
        const_group_numbers = constants['group_numbers']
        const_group_specialization_row = constants['group_specialization_row']
        const_group_specializations = constants['specialization']

        for column in range(const_first_group_column, const_last_group_column + 1):
            viewed_group_cell = worksheet.cell(
                row=const_group_row, column=column)
            viewed_group_value = viewed_group_cell.value
            index_of_group_name_and_specialization = column - 4
            if self.is_merged(worksheet, viewed_group_cell) == False:
                if viewed_group_value == None:
                    raise File_not_valid(
                        f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')
                result = re.fullmatch(r'Группа\s\d{3}', viewed_group_value)
                if result == None:
                    raise File_not_valid(
                        f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')
                if const_group_numbers[index_of_group_name_and_specialization] != viewed_group_value:
                    raise File_not_valid(
                        f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')
            else:
                try:
                    const_merged_group = constants['merged_group']
                    viewed_group_value = self.get_merged_cell_value(
                        worksheet, viewed_group_cell)
                    if viewed_group_value != const_merged_group:
                        raise File_not_valid(
                            f'Объединённая ячейка в {viewed_group_cell.coordinate} в листе {worksheet_name}')
                except:
                    raise File_not_valid(
                        f'Объединённая ячейка в {viewed_group_cell.coordinate} в листе {worksheet_name}')
            viewed_group_cell = worksheet.cell(
                row=const_group_specialization_row, column=column)
            if const_group_specializations[index_of_group_name_and_specialization] != viewed_group_cell.value:
                raise File_not_valid(
                    f'Ошибка в специализации в {viewed_group_cell.coordinate} в листе {worksheet_name}')

    def check_class_schedule(self, work_book, file_name):
        message = ''
        for worksheet_name in work_book.sheetnames:
            if worksheet_name.lower() in configurations.words_to_skip:
                continue
            if self.is_reason_to_skip(worksheet_name.lower()) == True:
                continue
            message += f'Лист {worksheet_name}\n'
            worksheet = work_book[worksheet_name]
            try:
                message += self.check_worksheet_schedule(worksheet, file_name)
            except:
                message += 'Критическая Ошибка'
            if 'Ошибка' in message:
                return message

        return message

    def check_worksheet_schedule(self, worksheet, file_name):
        message = ''
        constants = self.return_current_file_constants(file_name)
        const_first_class_cell_column = constants['first_group_column']
        const_last_class_cell_column = constants['last_group_column']
        const_first_class_cell_row = constants['first_time_row']
        error_values = []
        for row in range(const_first_class_cell_row, 50):
            for column in range(const_first_class_cell_column, const_last_class_cell_column + 1):
                viewed_class_cell = worksheet.cell(row=row, column=column)
                if self.is_merged(worksheet, viewed_class_cell):
                    viewed_class_value = self.get_merged_cell_value(
                        worksheet, viewed_class_cell)
                else:
                    viewed_class_value = viewed_class_cell.value
                if viewed_class_value == None:
                    continue
                cell_data = viewed_class_value.split('\n')
                if '' in cell_data or ' ' in cell_data:
                    message += f'Пустая строка в {viewed_class_cell.coordinate}\n'
                    continue
                if '  ' in cell_data[0]:
                    message += f'Минимум 2 пробела в {viewed_class_cell.coordinate}\n'
                    continue
                if cell_data[0] == 'Начальник УМУ Паульс А.А.':
                    if message == '':
                        message += 'Предметы ОК\n'
                    return message
                # print(viewed_class_cell.coordinate)
                # print(viewed_class_value)
                if 'ТиМ ИВС' in cell_data:
                    continue
                    # TODO а что если перед или после тире не будет пробела?
                    if len(cell_data) == 3:
                        subject = cell_data[0]
                        first_teacher = cell_data[1].split(' - ')[1]
                        second_teacher = cell_data[2].split(' - ')[1]
                        if subject not in configurations.existing_subjects:
                            message += f'Ошибка в предмете в  {viewed_class_cell.coordinate}\n'
                        if first_teacher not in configurations.existing_teachers:
                            message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                        if second_teacher not in configurations.existing_teachers:
                            message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                elif 'Элективные дисциплины по ФКиС' in cell_data:
                    continue
                    if len(cell_data) == 3:
                        subject = cell_data[0]
                        first_teacher = cell_data[1].split(' - ')[1]
                        second_teacher = cell_data[2].split(' - ')[1]
                        if subject not in configurations.existing_subjects:
                            message += f'Ошибка в предмете в  {viewed_class_cell.coordinate}\n'
                        if first_teacher not in configurations.existing_teachers:
                            message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                        if second_teacher not in configurations.existing_teachers:
                            message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                elif cell_data[0] in configurations.strings_to_skip_while_no_format:
                    continue
                elif len(cell_data) == 1:
                    continue
                elif len(cell_data) > 1:
                    if cell_data[1] in configurations.strings_to_skip_while_no_format:
                        continue
                elif 'практика' in cell_data[1]:
                    dates = cell_data[0]
                    practice = cell_data[1]
                    result = re.fullmatch(
                        r'С\s\d{2}[.]\d{2}[.]\sпо\s\d{2}[.]\d{2}[.]', dates)
                    if result == None:
                        message += f'Ошибка в датах в {viewed_class_cell.coordinate}\n'
                    if practice not in configurations.existing_practice:
                        message += f'Ошибка в практике в {viewed_class_cell.coordinate}\n'
                elif len(cell_data) == 4:
                    location = cell_data[0]
                    subject = cell_data[1]
                    type_of_subject = cell_data[2]
                    teacher = cell_data[3]
                    if 'ауд' in location:
                        result = re.fullmatch(r'ауд.\s\d{1,3}', location)
                        if result == None:
                            result_2 = re.fullmatch(
                                r'ауд.\s\d{1,3}\sИМиСТ', location)
                            if result_2 == None:
                                message += f'Ошибка в формате аудитории в  {viewed_class_cell.coordinate}\n'
                        auditorium = location[5:]
                        if auditorium not in configurations.existing_locations:
                            message += f'Ошибка в аудитории в  {viewed_class_cell.coordinate}\n'
                    if subject not in configurations.existing_subjects:
                        message += f'Ошибка в предмете в  {viewed_class_cell.coordinate}\n'
                    if type_of_subject not in configurations.existing_type_of_subject_type_1:
                        message += f'Ошибка в типе предмета в  {viewed_class_cell.coordinate}\n'
                    if teacher not in configurations.existing_teachers:
                        message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                elif len(cell_data) == 3:
                    # ловс 2 тут вопрос не решён
                    location = cell_data[0]
                    subject = cell_data[1]
                    teacher = cell_data[2]
                    if location not in configurations.existing_locations:
                        message += f'Ошибка в локации в  {viewed_class_cell.coordinate}\n'
                    if subject not in configurations.existing_subjects:
                        message += f'Ошибка в предмете в  {viewed_class_cell.coordinate}\n'
                    if teacher not in configurations.existing_teachers:
                        message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                elif len(cell_data) == 2:
                    subject = cell_data[0]
                    teacher = cell_data[1]
                    if subject in configurations.strings_to_skip_while_no_format:
                        continue
                    # if ' - ' in teacher:

                    if subject not in configurations.existing_subjects:
                        message += f'Ошибка в предмете в  {viewed_class_cell.coordinate}\n'
                    if teacher not in configurations.existing_teachers:
                        message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                if len(message) >= 2000:
                    return message
        if message == '':
            message += 'Предметы ОК'
        return message

    def return_current_file_constants(self, work_file_name):
        clear_file_name = self.find_clear_file_name(work_file_name)
        constants = configurations.group_constants[clear_file_name]
        return constants

    def check_file_name(self, work_file_name):
        if '1_zovs' in work_file_name:
            return 'Имя файла ОК\n'
        elif '2_zovs' in work_file_name:
            return 'Имя файла ОК\n'
        elif '3_zovs' in work_file_name:
            return 'Имя файла ОК\n'
        elif '4_zovs' in work_file_name:
            return 'Имя файла ОК\n'
        elif '1_lovs' in work_file_name:
            return 'Имя файла ОК\n'
        elif '2_lovs' in work_file_name:
            return 'Имя файла ОК\n'
        elif '3_lovs' in work_file_name:
            return 'Имя файла ОК\n'
        elif '4_lovs' in work_file_name:
            return 'Имя файла ОК\n'
        else:
            raise File_not_valid(f'Ошибка в имени файла {work_file_name}')

    def find_clear_file_name(self, work_file_name):
        if '1_zovs' in work_file_name:
            return 'zovs_1'
        elif '2_zovs' in work_file_name:
            return 'zovs_2'
        elif '3_zovs' in work_file_name:
            return 'zovs_3'
        elif '4_zovs' in work_file_name:
            return 'zovs_4'
        elif '1_lovs' in work_file_name:
            return 'lovs_1'
        elif '2_lovs' in work_file_name:
            return 'lovs_2'
        elif '3_lovs' in work_file_name:
            return 'lovs_3'
        elif '4_lovs' in work_file_name:
            return 'lovs_4'
        else:
            return None
