from openpyxl import load_workbook
import glob
import os

import sys
sys.path.append(os.path.join(os.path.abspath(os.path.dirname(__file__)), 'excel_validators'))

import db_funcs_for_subjects_db
import configurations
import excel_validator_imist
from excel_handler import Excel_handler


class Excel_parser_imist(Excel_handler):

    def run_parser(self, route):
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        validator = excel_validator_imist.Excel_validator_imist()

        for work_file_name in work_files:
            # validator.run_validator_for_excel_parser(route)
            db_name = self.return_db_name(work_file_name)
            work_book = load_workbook(work_file_name)
            constants = self.return_current_file_constants(work_file_name)

            self.create_db_for_parsing(work_book, constants, db_name)
            self.prepare_to_parse_work_file(work_book, constants, db_name)

    def prepare_to_parse_work_file(self, work_book, constants, db_name):
        number_of_groups = constants['number_of_groups']
        group_cell_constants = self.get_group_cell_constants(constants)
        groups_name = self.get_groups_name(constants)

        for worksheet_name in work_book.sheetnames:
            if self.is_reason_to_skip(worksheet_name) == True:
                continue
            worksheet = work_book[worksheet_name]

            for number in range(number_of_groups):
                first_row = group_cell_constants[number][0]
                last_row = group_cell_constants[number][1]
                first_column = group_cell_constants[number][2]
                date_column = constants['date_column']
                time_column = constants['time_column']
                group_name = groups_name[number]

                self.parse_worksheet(worksheet, db_name, group_name, date_column,
                                     time_column, first_row, last_row, first_column)

    def parse_worksheet(self, worksheet, db_name, group_name, date_column, time_column, first_lesson_cell_row, last_lesson_cell_row, first_lesson_cell_column):
        for row in range(first_lesson_cell_row, last_lesson_cell_row + 1):
            subject, subject_type = self.get_subject_and_subject_type(
                worksheet, row, first_lesson_cell_column)
            location = self.get_loction_or_teacher_value(
                worksheet, row, first_lesson_cell_column + 1)
            teacher = self.get_loction_or_teacher_value(
                worksheet, row, first_lesson_cell_column + 2)

            viewed_date_cell = worksheet.cell(row=row, column=date_column)
            viewed_date_cell_value = self.get_merged_cell_value(
                worksheet, viewed_date_cell)
            viewed_time_cell = worksheet.cell(row=row, column=time_column)
            viewed_time_cell_value = viewed_time_cell.value
            db_funcs_for_subjects_db.save_subj_imist(
                db_name, viewed_date_cell_value, viewed_time_cell_value, group_name, subject, subject_type, location, teacher)

    def get_loction_or_teacher_value(self, worksheet, row, column):
        viewed_cell = worksheet.cell(row=row, column=column)
        viewed_cell_value = viewed_cell.value
        if viewed_cell_value == None:
            viewed_cell_value = 'Нет предмета'
        return viewed_cell_value

    def get_subject_and_subject_type(self, worksheet, row, first_lesson_cell_column):
        viewed_lesson_cell = worksheet.cell(
            row=row, column=first_lesson_cell_column)
        viewed_lesson_cell_value = viewed_lesson_cell.value
        if viewed_lesson_cell_value == None:
            subject = 'Нет предмета'
            subject_type = 'Нет предмета'
        else:
            viewed_lesson_and_type_cell_value = viewed_lesson_cell_value.split(
                '\n')
            subject = viewed_lesson_and_type_cell_value[0]
            subject_type = viewed_lesson_and_type_cell_value[1]
        return subject, subject_type

    def create_db_for_parsing(self, work_book, constants, db_name):
        db_funcs_for_subjects_db.drop_db(db_name)
        db_funcs_for_subjects_db.create_db_imist(db_name)

        for worksheet_name in work_book.sheetnames:
            if self.is_reason_to_skip(worksheet_name) == True:
                continue
            worksheet = work_book[worksheet_name]
            self.create_dates_and_times_and_groups_in_db(
                worksheet, db_name, constants)

    def create_dates_and_times_and_groups_in_db(self, worksheet, db_name, constants):
        groups_name = self.get_groups_name(constants)

        date_column = constants['date_column']
        first_date_row = constants['first_date_row']
        last_date_row = constants['last_date_row']

        time_column = constants['time_column']

        for number in range(int(constants['number_of_groups'])):
            for row in range(first_date_row, last_date_row + 1):
                viewed_date_cell = worksheet.cell(row=row, column=date_column)
                viewed_date_value = self.get_merged_cell_value(
                    worksheet, viewed_date_cell)

                viewed_time_cell = worksheet.cell(row=row, column=time_column)
                viewed_time_value = viewed_time_cell.value

                group_name = groups_name[number]

                db_funcs_for_subjects_db.save_date_and_time_and_group_imist(
                    db_name, viewed_date_value, viewed_time_value, group_name)

    def get_group_cell_constants(self, constants):
        number_of_groups = constants['number_of_groups']

        group_cell_constants = [
            [
                constants['first_group_first_lesson_cell_row'],
                constants['first_group_last_lesson_cell_row'],
                constants['first_group_first_lesson_cell_column'],
            ],
            [
                constants['second_group_first_lesson_cell_row'],
                constants['second_group_last_lesson_cell_row'],
                constants['second_group_first_lesson_cell_column'],
            ],
            [
                constants['third_group_first_lesson_cell_row'],
                constants['third_group_last_lesson_cell_row'],
                constants['third_group_first_lesson_cell_column'],
            ],
        ]
        if number_of_groups == 4:
            fourth_group_first_lesson_cell_row = constants['fourth_group_first_lesson_cell_row']
            fourth_group_last_lesson_cell_row = constants['fourth_group_last_lesson_cell_row']
            fourth_group_first_lesson_cell_column = constants['fourth_group_first_lesson_cell_column']
            group_cell_constants.append(
                [fourth_group_first_lesson_cell_row, fourth_group_last_lesson_cell_row, fourth_group_first_lesson_cell_column])

        return group_cell_constants

    def get_groups_name(self, constants):
        number_of_groups = constants['number_of_groups']

        groups_name = [
            constants['first_group_number'],
            constants['second_group_number'],
            constants['third_group_number'],
        ]
        if number_of_groups == 4:
            groups_name.append(constants['fourth_group_number'])

        return groups_name

    def return_current_file_constants(self, work_file_name):
        clear_file_name = self.find_clear_file_name(work_file_name)
        constants = configurations.timetable_constants[clear_file_name]
        return constants

    def return_db_name(self, file_name):
        if '1_imist' in file_name:
            return 'imist_1'
        elif '2_imist' in file_name:
            return 'imist_2'
        elif '3_imist' in file_name:
            return 'imist_3'
        elif '4_imist' in file_name:
            return 'imist_4'

    def find_clear_file_name(self, file_name):
        if '1_imist' in file_name:
            return 'imist_1'
        elif '2_imist' in file_name:
            return 'imist_2'
        elif '3_imist' in file_name:
            return 'imist_3'
        elif '4_imist' in file_name:
            return 'imist_4'
        else:
            return None


parser = Excel_parser_imist()
if __name__ == "__main__":
    parser.run_parser('full_time_undergraduate_imist')
