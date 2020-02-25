import time
from openpyxl import Workbook, load_workbook, utils
import glob

import db_funcs_for_subjects_db


class Excel_parser():

    def is_merged(self, work_sheet, row, column):
        cell = work_sheet.cell(row, column)
        for mergedCell in work_sheet.merged_cells.ranges:
            if (cell.coordinate in mergedCell):
                return True
        return False
    
    def is_time(self, time):
        times = ['9:45', '09:45', '9.45', '09.45', '11:30', '11.30', '13:30', 
        '13.30', '15:15', '15.15', '17:00', '17.00', '18:40', '18.40','9:45:00', 
        '09:45:00', '9.45:00', '09.45:00', '11:30:00', '11.30:00', '13:30:00', 
        '13.30:00', '15:15:00','15.15:00', '17:00:00', '17.00:00', '18:40:00', '18.40:00']
        if time in times:
            return True
        else:
            return False

    def format_time(self, time):
        # в расписании чаще всего косячат тут
        first_times = ['9:45', '09:45', '9.45', '09.45', '9:45:00', '09:45:00', '9.45:00', '09.45:00']
        if time in first_times:
            time = '9:45'
        if '.' in time:
            time = time.replace('.', ':')
        if len(time) >= 8 and time[-1] == '0' and time[-2] == '0' and time[-3] == ':':
            if len(time) == 8:
                time = time[:5]
            elif len(time) == 7:
                time = time[:4]
        return time

    def save_dates_and_times(self, db_name, dates, times):
        for date in dates:
            for time in times:
                db_funcs_for_subjects_db.save_date_and_time(db_name, date, time)

    def format_dates(self, dates):
        list_of_dates = [element.replace(' ', '') for element in dates.rstrip().split('\n')]
        formatted_list_of_dates = []
        for date in list_of_dates:
            if date == '':
                continue
            elif date[-1] != '.':
                date = date + '.'
                formatted_list_of_dates.append(date)
            if len(date) >= 12:
                first_date = date[:6]
                second_date = date[6:]
                formatted_list_of_dates.append(first_date)
                formatted_list_of_dates.append(second_date)
                continue
            formatted_list_of_dates.append(date)
        return formatted_list_of_dates

class Excel_parser_magistracy_fk(Excel_parser):
    # константное значение
    # пока что во всех расписаниях колонка первых групп одинакова 
    const_first_group_column = 4
    const_time_column = 3
    const_dates_column = 1
    const_quantity_of_rows = 50


    const_magistracy_fk_1_kurs_groups = ['гр_1', 'гр_2', 'гр_3', 'гр_4', 'гр_5', 
    'гр_6', 'гр_7', 'гр_8', 'гр_9', 'гр_10', 'гр_11', 'гр_12', 'гр_13']

    mag_2_kurs_group_1 = "гимнастика_плавание_футбол"
    mag_2_kurs_group_2 = "лёгкая_атлетика_спортивные_игры"
    mag_2_kurs_group_3 = "велоспорт_водномоторный_и_парусный_спорт_гребной_спорт_атлетизм_бокс_борьба_фехтование_нвс"
    mag_2_kurs_group_4 = "хоккей_керлинг_биатлон_лыжный_спорт_конькобежный_спорт_и_фк"
    mag_2_kurs_group_5 = "профессиональное_образование_в_сфере_физической_культуры_и_спорта"
    mag_2_kurs_group_6 = "комплексное_научное_обеспечение_спортивной_подготовки"
    mag_2_kurs_group_7 = "физкультурно_оздоровительная_работа"
    mag_2_kurs_group_8 = "комплексная_реабилитация_в_физической_культуре_и_спорте"
    mag_2_kurs_group_9 = "медико_биологическое_сопровождение_физической_культуры_и_спорта"
    mag_2_kurs_group_10 = "руководитель_организации_в_области_физической_культуры_и_спорта"
    mag_2_kurs_group_11 = "подготовка_спортивных_сборных_команд_по_избранному_виду_спорта"
    mag_2_kurs_group_12 = "психология_спорта"

    const_magistracy_fk_2_kurs_groups = [mag_2_kurs_group_1, mag_2_kurs_group_2, 
        mag_2_kurs_group_3, mag_2_kurs_group_4, mag_2_kurs_group_5, mag_2_kurs_group_6,
        mag_2_kurs_group_7, mag_2_kurs_group_8, mag_2_kurs_group_9, mag_2_kurs_group_10,
        mag_2_kurs_group_11, mag_2_kurs_group_12]

    def run_parser(self, route):
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        print(work_files)
        for work_file in work_files:
            print(work_file)
            self.parse_work_file(work_file)

    def parse_work_file(self, work_file):
        db_name = self.return_db_name(work_file)
        db_funcs_for_subjects_db.drop_db(db_name)
        db_funcs_for_subjects_db.create_db(db_name)

        work_book = load_workbook(work_file)
        print(work_book)
        self.create_groups_in_db(work_book, db_name)

        for ws in work_book.sheetnames:
            work_sheet = work_book[ws]
            ws_date = str(work_sheet)
            month_to_skip = ['09', '10', '11', '12', '01']
            if ws_date[17:19] in month_to_skip:
                print('skipped' + ws_date)
                continue
            self.create_dates_and_times_in_db(work_sheet, db_name)
            self.self.parse_work_sheet(work_sheet, db_name)

    def parse_work_file_using_name(self, name, route):
        print('Парсер запущен на ' + name)
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        for work_file in work_files:
            if name in work_file:
                db_name = self.return_db_name(work_file)
                work_book = load_workbook(work_file)
                db_funcs_for_subjects_db.drop_db(db_name)
                db_funcs_for_subjects_db.create_db(db_name)
                self.create_groups_in_db(work_book, db_name)
                self.create_dates_and_times_in_db(work_book, db_name)

                for ws in work_book.sheetnames:
                    work_sheet = work_book[ws]
                    ws_date = str(work_sheet)
                    month_to_skip = ['09', '10', '11', '12', '01']
                    if ws_date[17:19] in month_to_skip:
                        print('skipped' + ws_date)
                        continue
                    self.self.parse_work_sheet(work_sheet, db_name)

    def return_db_name(self, file_name):
        if 'magistracy_fk_full_time_1_kurs' in file_name:
            return 'magistracy_fk_full_time_1_kurs'
        elif 'magistracy_fk_full_time_2_kurs' in file_name:
            return 'magistracy_fk_full_time_2_kurs'

    def create_groups_in_db(self, work_book, db_name):
        work_sheet = work_book[work_book.sheetnames[0]]
        first_group_name = ''
        if db_name == 'magistracy_fk_full_time_1_kurs':
            first_group_name = 'гр_1'
        elif db_name == 'magistracy_fk_full_time_2_kurs':
            first_group_name = 'гимнастика_плавание_футбол'
        list_of_groups = self.return_all_groups_names(work_sheet, first_group_name)
        db_funcs_for_subjects_db.save_groups(db_name, list_of_groups)

    def return_all_groups_names(self, work_sheet, first_group_name):
        groups_names = []
        row_number = self.find_number_of_groups_cell_row(work_sheet, first_group_name)   #   ПОЛИМОРФНАЯ ФУНКЦИЯ
        first_group_column = self.const_first_group_column
        for column in range(first_group_column, 25):
            group_cell = work_sheet.cell(row = row_number, column = column).value
            if type(group_cell) == str :
                group_cell = self.format_group_name_magistracy_fk_2_kurs(group_cell.lower().rstrip())
                if group_cell in self.const_magistracy_fk_1_kurs_groups or group_cell in self.const_magistracy_fk_2_kurs_groups:
                    groups_names.append(self.format_group_name_magistracy_fk(group_cell))
        return groups_names

    def format_group_name_magistracy_fk(self, group_name):  #    ПОЛИМОРФНАЯ?
        if ' ' in group_name:
            group_name = group_name.replace(' ', '_')
        if 'группа' in group_name:
            first = group_name[0]
            second = group_name[2:]
            group_name = second + '_' + first
        if 'гр.' in group_name and len(group_name) == 5:
            first = group_name[0]
            second = group_name[2:]
            group_name = second + '_' + first
        if 'гр.' in group_name and len(group_name) == 6:
            first = group_name[0:2]
            second = group_name[3:]
            group_name = second + '_' + first
        if '.' in group_name:
            group_name = group_name.replace('.', '')
        return group_name

    def format_group_name_magistracy_fk_2_kurs(self, group_name):
        group_name = group_name.lower()
        if '\n' in group_name:
            group_name = group_name.replace('\n', '')
        if '"' in group_name:
            group_name = group_name.replace('"', '')
        group_name = group_name.rstrip()
        if ';' in group_name:
            group_name = group_name.replace(';', '_')
        if ' ' in group_name:
            group_name = group_name.replace(' ', '_')
        if '__' in group_name:
            group_name = group_name.replace('__', '_')
        if group_name[0] == '_':
            group_name = group_name[1:]
        if group_name[-1] == '_':
            group_name = group_name[:-1]
        if '-' in group_name:
            group_name = group_name.replace('-', '_')
        if 'гр.' in group_name and len(group_name) == 5:
            first = group_name[0]
            second = group_name[2:]
            group_name = second + '_' + first
        if 'гр.' in group_name and len(group_name) == 6:
            first = group_name[0:2]
            second = group_name[3:]
            group_name = second + '_' + first
        if '.' in group_name:
            group_name = group_name.replace('.', '')
        return group_name

    def find_number_of_groups_cell_row(self, work_sheet, first_group_name):
        first_group_column = self.const_first_group_column
        for row in range(1, 10):
            viewed_cell = str(work_sheet.cell(row = row, column = first_group_column).value).rstrip()
            if type(viewed_cell) == str:
                viewed_cell = self.format_group_name_magistracy_fk_2_kurs(viewed_cell)
                if first_group_name in viewed_cell:
                    return row

    def get_dates(self, work_sheet, row, dates_column):
        dates = work_sheet.cell(row = row, column = dates_column).value
        if self.is_merged(work_sheet, row, dates_column):
            dates = self.get_value_of_merged_call(work_sheet, row, dates_column)
        dates = self.format_dates(dates)
        return dates    

    def get_value_of_merged_call(self, work_sheet, row, column):
        cell = work_sheet.cell(row, column)
        for mergedCell in work_sheet.merged_cells.ranges:
            if (cell.coordinate in mergedCell):
                column_coor = list(mergedCell)[0][1]
                row_coor = list(mergedCell)[1][1]
                result_value = work_sheet.cell(row = row_coor, column = column_coor).value
                return result_value

    def create_dates_and_times_in_db(self, work_sheet, db_name):
        dates_column = self.const_dates_column
        time_column = self.const_time_column

        first_row = self.find_row_of_first_lesson(work_sheet)
        times = []
        for row in range(1, self.const_quantity_of_rows):
            time_cell = str(work_sheet.cell(row = row, column = time_column).value)
            if time_cell != None and self.is_time(time_cell):
                time_value = self.format_time(time_cell)
                if time_value == '9:45':
                    times = []
                    times.append(time_value)
                    dates = self.get_dates(work_sheet, row, dates_column)
                elif time_value == '11:30' or time_value == '13:30' or time_value == '15:15':
                    times.append(time_value)
                elif time_value == '17:00':
                    times.append(time_value)
                    next_cell = str(work_sheet.cell(row = row + 1, column = time_column).value)
                    next_cell = self.format_time(next_cell) if self.is_time(next_cell) else False

                    after_next_cell = str(work_sheet.cell(row = row + 2, column = time_column).value)
                    after_next_cell = self.format_time(after_next_cell) if self.is_time(after_next_cell) else False

                    if self.is_time(next_cell) or self.is_time(after_next_cell):
                        if next_cell == '9:45' or after_next_cell == '9:45' and next_cell != '18:40':
                            self.save_dates_and_times(db_name, dates, times)
                            continue                               
                        elif next_cell == '18:40' or after_next_cell == '18:40':
                            continue
                    elif next_cell == False and after_next_cell == False:
                        self.save_dates_and_times(db_name, dates, times)
                elif time_value == '18:40':
                    times.append(time_value)
                    dates = self.get_dates(work_sheet, row, dates_column)
                    self.save_dates_and_times(db_name, dates, times)


    def find_row_of_first_lesson(self, work_sheet):
        time_column = self.const_time_column
        times = ['9:45', '09:45', '9.45', '09.45', '9:45:00', '09:45:00', '9.45:00', '09.45:00']
        for row in range(1, 10):
            time_cell = str(work_sheet.cell(row = row, column = time_column).value)
            if time_cell in times:
                return row 

    def parse_work_sheet(self, work_sheet, db_name):
        time_column = self.const_time_column
        dates_column = self.const_dates_column

        if db_name == 'magistracy_fk_full_time_1_kurs':
            first_group_name = 'гр_1'
        elif db_name == 'magistracy_fk_full_time_2_kurs':
            first_group_name = 'гимнастика_плавание_футбол'

        groups_columns = self.return_columns_numbers_of_all_groups_cells(work_sheet, first_group_name)
        first_row = self.find_row_of_first_lesson(work_sheet)
        groups_row = self.find_number_of_groups_cell_row(work_sheet, first_group_name)
        times = ['9:45', '11:30', '13:30', '15:15', '17:00', '18:40','9:45:00', 
        '09:45:00', '9.45:00', '09.45:00', '11:30:00', '11.30:00', '13:30:00', 
        '13.30:00', '15:15:00','15.15:00', '17:00:00', '17.00:00', '18:40:00', '18.40:00']
        print('ws start ' + str(work_sheet))
        for column in groups_columns:
            for row in range(first_row, self.const_quantity_of_rows):
                subject = work_sheet.cell(row = row, column = column).value
                if self.is_merged(work_sheet, row, column):
                    subject = self.get_value_of_merged_call(work_sheet, row, column)
                if subject == None:
                    subject = 'нет предмета'
                time_cell = work_sheet.cell(row = row, column = time_column).value
                if time_cell == None:
                    continue
                time = self.format_time(str(time_cell))
                if time in times:
                    dates = work_sheet.cell(row = row, column = dates_column).value
                    if self.is_merged(work_sheet, row, dates_column):
                        dates = self.get_value_of_merged_call(work_sheet, row, dates_column)
                    group_name = work_sheet.cell(row = groups_row, column = column).value.rstrip()
                    group_name = self.format_group_name_magistracy_fk_2_kurs(group_name)
                    self.save_subj_in_db(db_name, dates, time, group_name, subject)
                else:
                    print(subject)
        print('ws finished')

    def save_subj_in_db(self, db_name, dates, time, group_name, subject):
        dates = self.format_dates(dates)
        time = self.format_time(time)
        group_name = self.format_group_name_magistracy_fk(group_name.lower())
        for date in dates:
            db_funcs_for_subjects_db.save_subj(db_name, date, time, group_name, subject)


    def return_columns_numbers_of_all_groups_cells(self, work_sheet, first_group_name):
        columns_numbers_of_all_groups_cells = []
        row_number = self.find_number_of_groups_cell_row(work_sheet, first_group_name)
        first_group_column = self.const_first_group_column
        for column in range(first_group_column, 25):
            group_cell = work_sheet.cell(row = row_number, column = column).value
            if type(group_cell) == str:
                group_cell = self.format_group_name_magistracy_fk_2_kurs(group_cell)
                if group_cell in self.const_magistracy_fk_1_kurs_groups or group_cell in self.const_magistracy_fk_2_kurs_groups:
                    columns_numbers_of_all_groups_cells.append(column)
        return columns_numbers_of_all_groups_cells


class Excel_parser_magistracy_imist(Excel_parser):
    # константное значение
    # пока что во всех расписаниях колонка первых групп одинакова 
    const_first_group_column = 4
    const_time_column = 3
    const_dates_column = 1
    const_quantity_of_rows = 50

    group_1_kurs_1 = "менеджмент_направленность_профиль_менеджмент_в_спорте"
    group_2_kurs_1 = "туризм_направленность_профиль_туристская_деятельность_с_сфере_физической_культуры_и_спорта"
    group_3_kurs_1 = "журналистика_направленность_профиль_спортивная_журналистика"
                        
    group_1_kurs_2 = "государственное_и_муниципальное_управление_направленность_профиль_государственное_и_муниципальное_управление_в_отрасли_физической_культуры_и_спорта"
    group_2_kurs_2 = "менеджмент_направленность_профиль_менеджмент_в_спорте"
    group_3_kurs_2 = "туризм_направленность_профиль_туристская_деятельность_с_сфере_физической_культуры_и_спорта"
    group_4_kurs_2 = "журналистика_направленность_профиль_спортивная_журналистика"

    const_magistracy_imst_groups = [group_1_kurs_1, group_2_kurs_1, group_3_kurs_1,
        group_1_kurs_2, group_2_kurs_2, group_3_kurs_2, group_4_kurs_2]

    def run_parser(self, route):
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        print(work_files)
        for work_file in work_files:
            print(work_file)
            if '1_kurs' in str(work_file):
                number_of_kurs = 1
            elif '2_kurs' in str(work_file):
                number_of_kurs = 2
            self.parse_work_file(work_file, number_of_kurs)


    def parse_work_file(self, work_file, number_of_kurs):
        db_name = self.return_db_name(work_file)
        db_funcs_for_subjects_db.drop_db(db_name)
        db_funcs_for_subjects_db.create_db(db_name)

        work_book = load_workbook(work_file)
        self.create_groups_in_db(work_book, db_name, number_of_kurs)

        for ws in work_book.sheetnames:
            work_sheet = work_book[ws]
            ws_date = str(work_sheet)
            month_to_skip = ['09', '10', '11', '12', '01']
            if ws_date[17:19] in month_to_skip:
                print('skipped' + ws_date)
                continue
            self.create_dates_and_times_in_db(work_sheet, db_name)
            self.parse_work_sheet(work_sheet, db_name, number_of_kurs)

    def parse_work_file_using_name(self, name, route):
        print('Парсер запущен на ' + name)
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        for work_file in work_files:
            if name in work_file:
                db_name = self.return_db_name(work_file)
                work_book = load_workbook(work_file)
                db_funcs_for_subjects_db.drop_db(db_name)
                db_funcs_for_subjects_db.create_db(db_name)
                self.create_groups_in_db(work_book, db_name)
                self.create_dates_and_times_in_db(work_book, db_name)

                for ws in work_book.sheetnames:
                    work_sheet = work_book[ws]
                    ws_date = str(work_sheet)
                    month_to_skip = ['09', '10', '11', '12', '01']
                    if ws_date[17:19] in month_to_skip:
                        print('skipped' + ws_date)
                        continue
                    self.parse_work_sheet(work_sheet, db_name)

    def return_db_name(self, file_name):
        if 'magistracy_imst_full_time_1_kurs' in file_name:
            return 'magistracy_imist_full_time_1_kurs'
        elif 'magistracy_imst_full_time_2_kurs' in file_name:
            return 'magistracy_imist_full_time_2_kurs'

    def create_groups_in_db(self, work_book, db_name, number_of_kurs):
        work_sheet = work_book[work_book.sheetnames[0]]
        if number_of_kurs == 1:
            first_group_name = self.group_1_kurs_1
        elif number_of_kurs == 2:
            first_group_name = self.group_1_kurs_2
        list_of_groups = self.return_all_groups_names(work_sheet, first_group_name)
        db_funcs_for_subjects_db.save_groups(db_name, list_of_groups)

    def return_all_groups_names(self, work_sheet, first_group_name):
        groups_names = []
        row_number = self.find_number_of_groups_cell_row(work_sheet, first_group_name)
        first_group_column = self.const_first_group_column
        for column in range(first_group_column, 25):
            group_cell = work_sheet.cell(row = row_number, column = column).value
            if type(group_cell) == str :
                group_cell = group_cell.lower().rstrip()
                if type(group_cell) == str:
                    group_cell = self.format_group_name_magistracy_imst(group_cell)
                if group_cell in self.const_magistracy_imst_groups:
                    groups_names.append(group_cell)
        return groups_names

    def format_group_name_magistracy_imst(self, group_name):
        group_name = group_name[9:]
        if '\n' in group_name:
            group_name = group_name.replace('\n', ' ')
        group_name = group_name.rstrip()
        group_name = group_name.lower()
        if '                                                                                                                                                 ' in group_name:
            group_name = group_name.replace('                                                                                                                                                  ', ' ')
        if '      ' in group_name:
            group_name = group_name.replace('      ', ' ')
        if '                         ' in group_name:
            group_name = group_name.replace('                         ', '')
        if '  ' in group_name:
            group_name = group_name.replace('  ', ' ')
        if '  ' in group_name:
            group_name = group_name.replace('  ', ' ')
        if group_name[0] == '-':
            group_name = group_name[1:]
        if group_name[0] == ' ':
            group_name = group_name[1:]
        if ':' in group_name:
            group_name = group_name.replace(':', '')
        if ' ' in group_name:
            group_name = group_name.replace(' ', '_')
        if '(' in group_name:
            group_name = group_name.replace('(', '')
            group_name = group_name.replace(')', '')
        if '"' in group_name:
            group_name = group_name.replace('"', '')
        return group_name

    def find_number_of_groups_cell_row(self, work_sheet, first_group_name):
        first_group_column = self.const_first_group_column
        for row in range(1, 10):
            viewed_cell = str(work_sheet.cell(row = row, column = first_group_column).value).rstrip()
            if type(viewed_cell) == str:
                try:
                    viewed_cell = self.format_group_name_magistracy_imst(viewed_cell)
                except:
                    pass
            if type(viewed_cell) == str and first_group_name in viewed_cell:
                return row

    def get_dates(self, work_sheet, row, dates_column):
        dates = work_sheet.cell(row = row, column = dates_column).value
        if self.is_merged(work_sheet, row, dates_column):
            dates = self.get_value_of_merged_call(work_sheet, row, dates_column)
        dates = self.format_dates(dates)
        return dates    

    def get_value_of_merged_call(self, work_sheet, row, column):
        cell = work_sheet.cell(row, column)
        for mergedCell in work_sheet.merged_cells.ranges:
            if (cell.coordinate in mergedCell):
                column_coor = list(mergedCell)[0][1]
                row_coor = list(mergedCell)[1][1]
                result_value = work_sheet.cell(row = row_coor, column = column_coor).value
                return result_value

    def create_dates_and_times_in_db(self, work_sheet, db_name):
        dates_column = self.const_dates_column
        time_column = self.const_time_column

        first_row = self.find_row_of_first_lesson(work_sheet)
        times = []
        for row in range(1, self.const_quantity_of_rows):
            time_cell = str(work_sheet.cell(row = row, column = time_column).value)
            if time_cell != None and self.is_time(time_cell):
                time_value = self.format_time(time_cell)
                if time_value == '9:45':
                    times = []
                    times.append(time_value)
                    dates = self.get_dates(work_sheet, row, dates_column)
                elif time_value == '11:30' or time_value == '13:30' or time_value == '15:15':
                    times.append(time_value)
                elif time_value == '17:00':
                    times.append(time_value)
                    next_cell = str(work_sheet.cell(row = row + 1, column = time_column).value)
                    next_cell = self.format_time(next_cell) if self.is_time(next_cell) else False

                    after_next_cell = str(work_sheet.cell(row = row + 2, column = time_column).value)
                    after_next_cell = self.format_time(after_next_cell) if self.is_time(after_next_cell) else False

                    if self.is_time(next_cell) or self.is_time(after_next_cell):
                        if next_cell == '9:45' or after_next_cell == '9:45' and next_cell != '18:40':
                            self.save_dates_and_times(db_name, dates, times)
                            continue                               
                        elif next_cell == '18:40' or after_next_cell == '18:40':
                            continue
                    elif next_cell == False and after_next_cell == False:
                        self.save_dates_and_times(db_name, dates, times)
                elif time_value == '18:40':
                    times.append(time_value)
                    dates = self.get_dates(work_sheet, row, dates_column)
                    self.save_dates_and_times(db_name, dates, times)



    def find_row_of_first_lesson(self, work_sheet):
        time_column = self.const_time_column
        times = ['9:45', '09:45', '9.45', '09.45', '9:45:00', '09:45:00', '9.45:00', '09.45:00']
        for row in range(1, 10):
            time_cell = work_sheet.cell(row = row, column = time_column).value
            if time_cell in times:
                return row 

    def parse_work_sheet(self, work_sheet, db_name, number_of_kurs):
        time_column = self.const_time_column
        dates_column = self.const_dates_column

        if number_of_kurs == 1:
            first_group_name = self.group_1_kurs_1
        elif number_of_kurs == 2:
            first_group_name = self.group_1_kurs_2
        
        groups_column = self.return_columns_numbers_of_all_groups_cells(work_sheet, first_group_name)
        first_row = self.find_row_of_first_lesson(work_sheet)
        if first_row == None:
            return False
        groups_row = self.find_number_of_groups_cell_row(work_sheet, first_group_name)
        times = ['9:45', '11:30', '13:30', '15:15', '17:00', '18:40','9:45:00', 
        '09:45:00', '9.45:00', '09.45:00', '11:30:00', '11.30:00', '13:30:00', 
        '13.30:00', '15:15:00','15.15:00', '17:00:00', '17.00:00', '18:40:00', '18.40:00']
        print('ws start ' + str(work_sheet))
        for column in groups_column:
            for row in range(first_row, self.const_quantity_of_rows):
                subject = work_sheet.cell(row = row, column = column).value
                if self.is_merged(work_sheet, row, column):
                    subject = self.get_value_of_merged_call(work_sheet, row, column)
                if subject == None:
                    subject = 'нет предмета'
                time_cell = work_sheet.cell(row = row, column = time_column).value
                if time_cell == None:
                    continue
                time = self.format_time(str(time_cell))
                if time in times:
                    dates = work_sheet.cell(row = row, column = dates_column).value
                    if self.is_merged(work_sheet, row, dates_column):
                        dates = self.get_value_of_merged_call(work_sheet, row, dates_column)
                    group_name = work_sheet.cell(row = groups_row, column = column).value.rstrip()
                    self.save_subj_in_db(db_name, dates, time, group_name, subject)
                else:
                    print(subject)
        print('ws finished')

    def save_subj_in_db(self, db_name, dates, time, group_name, subject):
        dates = self.format_dates(dates)
        time = self.format_time(time)
        group_name = self.format_group_name_magistracy_imst(group_name.lower())
        for date in dates:
            db_funcs_for_subjects_db.save_subj(db_name, date, time, group_name, subject)

    def return_columns_numbers_of_all_groups_cells(self, work_sheet, first_group_name):
        columns_numbers_of_all_groups_cells = []
        row_number = self.find_number_of_groups_cell_row(work_sheet, first_group_name)
        first_group_column = self.const_first_group_column
        for column in range(first_group_column, 25):
            group_cell = work_sheet.cell(row = row_number, column = column).value
            if type(group_cell) == str:
                group_cell = self.format_group_name_magistracy_imst(group_cell.lower().rstrip())
                if '\n' in group_cell:
                    group_cell = group_cell.replace('\n', '')
                if group_cell[-1] == ' ':
                    group_cell = group_cell[:-1]
                if group_cell in self.const_magistracy_imst_groups:
                    columns_numbers_of_all_groups_cells.append(column)
        return columns_numbers_of_all_groups_cells

class Excel_parser_magistracy_afk(Excel_parser):
    # константное значение
    # пока что во всех расписаниях колонка первых групп одинакова 
    const_first_group_column = 4
    const_time_column = 3
    const_dates_column = 1
    const_quantity_of_rows = 50

    group_1_kurs_1 = "адаптивное_физическое_воспитание_в_системе_образования_обучающихся_с_овз"
    group_2_kurs_1 = "спортивная_подготовка_лиц_с_овз_включая_инвалидов"
    group_3_kurs_1 = "физическая_реабилитация"
    group_4_kurs_1 = "педагогическая_гидрореабилитация"
    group_5_kurs_1 = "современная_хореография_и_танцы_в_коррекции_нарушений_у_лиц_с_овз"

    group_1_kurs_2 = "адаптивное_физическое_воспитание_в_системе_образования_обучающихся_с_ограниченными_возможностями_здоровья"
    group_2_kurs_2 = "технологии_профилактики_и_коррекции_аддиктивного_поведения"
    group_3_kurs_2 = "спортивная_подготовка_лиц_с_ограниченными_возможностями_здоровья_включая_инвалидов"
    group_4_kurs_2 = "педагогическая_гидрореабилитация"
    group_5_kurs_2 = "физическая_реабилитация"
    group_6_kurs_2 = "современная_хореография_и_танцы_в_коррекции_нарушений_у_лиц_с_отклонениями_в_состоянии_здоровья"

    const_magistracy_imst_groups = [group_1_kurs_1, group_2_kurs_1, group_3_kurs_1,
        group_4_kurs_1, group_5_kurs_1, group_1_kurs_2, group_2_kurs_2, group_3_kurs_2, 
        group_4_kurs_2, group_5_kurs_2, group_6_kurs_2]

    def run_parser(self, route):
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        print(work_files)
        for work_file in work_files:
            print(work_file)
            if '1_kurs' in str(work_file):
                number_of_kurs = 1
            elif '2_kurs' in str(work_file):
                number_of_kurs = 2
            self.parse_work_file(work_file, number_of_kurs)


    def parse_work_file(self, work_file, number_of_kurs):
        db_name = self.return_db_name(work_file)
        print(work_file)
        db_funcs_for_subjects_db.drop_db(db_name)
        db_funcs_for_subjects_db.create_db(db_name)

        work_book = load_workbook(work_file)
        self.create_groups_in_db(work_book, db_name, number_of_kurs)

        for ws in work_book.sheetnames:
            work_sheet = work_book[ws]
            ws_date = str(work_sheet)
            month_to_skip = ['09', '10']#, '11', '12', '01']
            if ws_date[17:19] in month_to_skip:
                print('skipped' + ws_date)
                continue
            self.create_dates_and_times_in_db(work_sheet, db_name)
            self.parse_work_sheet(work_sheet, db_name, number_of_kurs)

    def parse_work_file_using_name(self, name, route):
        print('Парсер запущен на ' + name)
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        for work_file in work_files:
            if name in work_file:
                db_name = self.return_db_name(work_file)
                work_book = load_workbook(work_file)
                db_funcs_for_subjects_db.drop_db(db_name)
                db_funcs_for_subjects_db.create_db(db_name)
                self.create_groups_in_db(work_book, db_name)
                self.create_dates_and_times_in_db(work_book, db_name)

                for ws in work_book.sheetnames:
                    work_sheet = work_book[ws]
                    ws_date = str(work_sheet)
                    month_to_skip = ['09', '10']#, '11', '12', '01']
                    if ws_date[17:19] in month_to_skip:
                        print('skipped' + ws_date)
                        continue
                    self.parse_work_sheet(work_sheet, db_name)

    def return_db_name(self, file_name):
        if 'magistracy_afk_full_time_1_kurs' in file_name:
            return 'magistracy_afk_full_time_1_kurs'
        elif 'magistracy_afk_full_time_2_kurs' in file_name:
            return 'magistracy_afk_full_time_2_kurs'

    def create_groups_in_db(self, work_book, db_name, number_of_kurs):
        work_sheet = work_book[work_book.sheetnames[0]]
        if number_of_kurs == 1:
            first_group_name = self.group_1_kurs_1
        elif number_of_kurs == 2:
            first_group_name = self.group_1_kurs_2
        list_of_groups = self.return_all_groups_names(work_sheet, first_group_name)
        db_funcs_for_subjects_db.save_groups(db_name, list_of_groups)

    def return_all_groups_names(self, work_sheet, first_group_name):
        groups_names = []
        row_number = self.find_number_of_groups_cell_row(work_sheet, first_group_name)
        first_group_column = self.const_first_group_column
        for column in range(first_group_column, 25):
            group_cell = work_sheet.cell(row = row_number, column = column).value
            if type(group_cell) == str :
                group_cell = group_cell.lower().rstrip()
                if type(group_cell) == str:
                    group_cell = self.format_group_name_magistracy_afk(group_cell)
                if group_cell in self.const_magistracy_imst_groups:
                    groups_names.append(group_cell)
        return groups_names

    def format_group_name_magistracy_afk(self, group_name):
        group_name = group_name.lower()
        if '\n' in group_name:
            group_name = group_name.replace('\n', '')
        group_name = group_name.rstrip()
        if ' ' in group_name:
            group_name = group_name.replace(' ', '_')
        if ',' in group_name:
            group_name = group_name.replace(',', '')
        if group_name[0] == '_':
            group_name = group_name[1:]
        if '"' in group_name:
            group_name = group_name.replace('"', '')
        return group_name

    def find_number_of_groups_cell_row(self, work_sheet, first_group_name):
        first_group_column = self.const_first_group_column
        for row in range(1, 10):
            viewed_cell = str(work_sheet.cell(row = row, column = first_group_column).value).rstrip()
            if type(viewed_cell) == str:
                try:
                    viewed_cell = self.format_group_name_magistracy_afk(viewed_cell)
                except:
                    pass
            if type(viewed_cell) == str and first_group_name in viewed_cell:
                return row

    def get_dates(self, work_sheet, row, dates_column):
        dates = work_sheet.cell(row = row, column = dates_column).value
        if self.is_merged(work_sheet, row, dates_column):
            dates = self.get_value_of_merged_call(work_sheet, row, dates_column)
        dates = self.format_dates(dates)
        return dates    

    def get_value_of_merged_call(self, work_sheet, row, column):
        cell = work_sheet.cell(row, column)
        for mergedCell in work_sheet.merged_cells.ranges:
            if (cell.coordinate in mergedCell):
                column_coor = list(mergedCell)[0][1]
                row_coor = list(mergedCell)[1][1]
                result_value = work_sheet.cell(row = row_coor, column = column_coor).value
                return result_value

    def create_dates_and_times_in_db(self, work_sheet, db_name):
        dates_column = self.const_dates_column
        time_column = self.const_time_column

        first_row = self.find_row_of_first_lesson(work_sheet)
        times = []
        for row in range(1, self.const_quantity_of_rows):
            time_cell = str(work_sheet.cell(row = row, column = time_column).value)
            if time_cell != None and self.is_time(time_cell):
                time_value = self.format_time(time_cell)
                if time_value == '9:45':
                    times = []
                    times.append(time_value)
                    dates = self.get_dates(work_sheet, row, dates_column)
                elif time_value == '11:30' or time_value == '13:30' or time_value == '15:15':
                    times.append(time_value)
                elif time_value == '17:00':
                    times.append(time_value)
                    next_cell = str(work_sheet.cell(row = row + 1, column = time_column).value)
                    next_cell = self.format_time(next_cell) if self.is_time(next_cell) else False

                    after_next_cell = str(work_sheet.cell(row = row + 2, column = time_column).value)
                    after_next_cell = self.format_time(after_next_cell) if self.is_time(after_next_cell) else False

                    if self.is_time(next_cell) or self.is_time(after_next_cell):
                        if next_cell == '9:45' or after_next_cell == '9:45' and next_cell != '18:40':
                            self.save_dates_and_times(db_name, dates, times)
                            continue                               
                        elif next_cell == '18:40' or after_next_cell == '18:40':
                            continue
                    elif next_cell == False and after_next_cell == False:
                        self.save_dates_and_times(db_name, dates, times)
                elif time_value == '18:40':
                    times.append(time_value)
                    dates = self.get_dates(work_sheet, row, dates_column)
                    self.save_dates_and_times(db_name, dates, times)



    def find_row_of_first_lesson(self, work_sheet):
        time_column = self.const_time_column
        times = ['9:45', '09:45', '9.45', '09.45', '9:45:00', '09:45:00', '9.45:00', '09.45:00']
        for row in range(1, 10):
            time_cell = work_sheet.cell(row = row, column = time_column).value
            if time_cell in times:
                return row 

    def parse_work_sheet(self, work_sheet, db_name, number_of_kurs):
        time_column = self.const_time_column
        dates_column = self.const_dates_column

        if number_of_kurs == 1:
            first_group_name = self.group_1_kurs_1
        elif number_of_kurs == 2:
            first_group_name = self.group_1_kurs_2
        
        groups_column = self.return_columns_numbers_of_all_groups_cells(work_sheet, first_group_name)
        first_row = self.find_row_of_first_lesson(work_sheet)
        if first_row == None:
            return False
        groups_row = self.find_number_of_groups_cell_row(work_sheet, first_group_name)
        times = ['9:45', '11:30', '13:30', '15:15', '17:00', '18:40','9:45:00', 
        '09:45:00', '9.45:00', '09.45:00', '11:30:00', '11.30:00', '13:30:00', 
        '13.30:00', '15:15:00','15.15:00', '17:00:00', '17.00:00', '18:40:00', '18.40:00']
        print('ws start ' + str(work_sheet))
        for column in groups_column:
            for row in range(first_row, self.const_quantity_of_rows):
                subject = work_sheet.cell(row = row, column = column).value
                if self.is_merged(work_sheet, row, column):
                    subject = self.get_value_of_merged_call(work_sheet, row, column)
                if subject == None:
                    subject = 'нет предмета'
                time_cell = work_sheet.cell(row = row, column = time_column).value
                if time_cell == None:
                    continue
                time = self.format_time(str(time_cell))
                if time in times:
                    dates = work_sheet.cell(row = row, column = dates_column).value
                    if self.is_merged(work_sheet, row, dates_column):
                        dates = self.get_value_of_merged_call(work_sheet, row, dates_column)
                    group_name = work_sheet.cell(row = groups_row, column = column).value.rstrip()
                    self.save_subj_in_db(db_name, dates, time, group_name, subject)
                else:
                    print(subject)
        print('ws finished')

    def save_subj_in_db(self, db_name, dates, time, group_name, subject):
        dates = self.format_dates(dates)
        time = self.format_time(time)
        group_name = self.format_group_name_magistracy_afk(group_name.lower())
        for date in dates:
            db_funcs_for_subjects_db.save_subj(db_name, date, time, group_name, subject)

    def return_columns_numbers_of_all_groups_cells(self, work_sheet, first_group_name):
        columns_numbers_of_all_groups_cells = []
        row_number = self.find_number_of_groups_cell_row(work_sheet, first_group_name)
        first_group_column = self.const_first_group_column
        for column in range(first_group_column, 25):
            group_cell = work_sheet.cell(row = row_number, column = column).value
            if type(group_cell) == str:
                group_cell = self.format_group_name_magistracy_afk(group_cell.lower().rstrip())
                if '\n' in group_cell:
                    group_cell = group_cell.replace('\n', '')
                if group_cell[-1] == ' ':
                    group_cell = group_cell[:-1]
                if group_cell in self.const_magistracy_imst_groups:
                    columns_numbers_of_all_groups_cells.append(column)
        return columns_numbers_of_all_groups_cells

class Excel_parser_undergraduate_imst(Excel_parser):
    # константное значение
    # пока что во всех расписаниях колонка первых групп одинакова 
    const_first_group_column = 7
    const_time_column = 6
    const_dates_column = 4
    const_quantity_of_rows = 50

    const_imist_groups = ['менеджмент', 'международные отношения', 'журналистика', 
        'туризм', 'сервис', 'реклама и связи с общественностью']

    const_imist_groups_with_numbers = ['менеджмент 38.03.02', 
        'международные отношения 41.03.05', 'журналистика 42.03.02',
        'сервис 43.03.01', 'р еклама и связи с общественностью 42.03.01', 'туризм 43.03.02']

    def run_parser(self, route):
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        print(work_files)
        for work_file in work_files:
            print(work_file)
            self.parse_work_file(work_file)

    def find_number_of_groups_cell_row(self, work_sheet):
        first_group_column = self.const_first_group_column
        for row in range(1, 10):
            viewed_cell = work_sheet.cell(row = row, column = first_group_column).value 
            if type(viewed_cell) == str and 'Менеджмент' in viewed_cell:
                return row

    def return_columns_numbers_of_all_groups_cells(self, work_sheet):
        columns_numbers_of_all_groups_cells = []
        row_number = self.find_number_of_groups_cell_row(work_sheet)
        first_group_column = self.const_first_group_column
        for column in range(first_group_column, 25):
            group_cell = work_sheet.cell(row = row_number, column = column).value
            if type(group_cell) == str:
                group_cell = group_cell.lower()
                if '\n' in group_cell:
                    group_cell = group_cell.replace('\n', '')
                if group_cell[-1] == ' ':
                    group_cell = group_cell[:-1]
                if group_cell in self.const_imist_groups_with_numbers:
                    columns_numbers_of_all_groups_cells.append(column)
        return columns_numbers_of_all_groups_cells

    def return_all_groups_names(self, work_sheet):
        groups_names = []
        row_number = self.find_number_of_groups_cell_row(work_sheet)
        first_group_column = self.const_first_group_column
        for column in range(first_group_column, 25):
            group_cell = work_sheet.cell(row = row_number, column = column).value
            if type(group_cell) == str :
                group_cell = group_cell.lower()
                for group in self.const_imist_groups:
                    if group in group_cell:
                        groups_names.append(self.format_group_name(group))
        return groups_names


    def find_row_of_first_lesson(self, work_sheet):
        time_column = self.const_time_column
        times = ['9:45', '09:45', '9.45', '09.45', '9:45:00', '09:45:00', '9.45:00', '09.45:00']

        for row in range(1, 10):
            time_cell = str(work_sheet.cell(row = row, column = time_column).value)
            if time_cell in times:
                return row 

    def get_value_of_merged_call(self, work_sheet, row, column):
        cell = work_sheet.cell(row, column)
        for mergedCell in work_sheet.merged_cells.ranges:
            if (cell.coordinate in mergedCell):
                column_coor = list(mergedCell)[0][1]
                row_coor = list(mergedCell)[1][1]
                result_value = work_sheet.cell(row = row_coor, column = column_coor).value
                return result_value

    def format_group_name(self, group_name):
        if 'менеджмент' in group_name:
            return 'менеджмент'
        elif 'международные отношения' in group_name:
            return 'международные_отношения'
        elif 'журналистика' in group_name:
            return 'журналистика'
        elif 'туризм' in group_name:
            return 'туризм'
        elif 'сервис' in group_name:
            return 'сервис'
        elif 'реклама и связи с общественностью' in group_name:
            return 'реклама_и_связи_с_общественностью'



    def save_subj_in_db(self, db_name, dates, time, group_name, subject):
        dates = self.format_dates(dates)
        time = self.format_time(time)
        group_name = self.format_group_name(group_name.lower())

        for date in dates:
            db_funcs_for_subjects_db.save_subj(db_name, date, time, group_name, subject)

    def parse_work_sheet(self, work_sheet, db_name):
        time_column = self.const_time_column
        dates_column = self.const_dates_column

        groups_column = self.return_columns_numbers_of_all_groups_cells(work_sheet)
        first_row = self.find_row_of_first_lesson(work_sheet)
        if first_row == None:
            return False
        groups_row = self.find_number_of_groups_cell_row(work_sheet)

        times = ['9:45', '11:30', '13:30', '15:15', '17:00', '18:40','9:45:00', 
        '09:45:00', '9.45:00', '09.45:00', '11:30:00', '11.30:00', '13:30:00', 
        '13.30:00', '15:15:00','15.15:00', '17:00:00', '17.00:00', '18:40:00', '18.40:00']
        print('ws start')
        for column in groups_column:
            for row in range(first_row, self.const_quantity_of_rows):
                subject = work_sheet.cell(row = row, column = column).value
                if self.is_merged(work_sheet, row, column):
                    subject = self.get_value_of_merged_call(work_sheet, row, column)
                if subject == None:
                    subject = 'нет предмета'
                time_cell = work_sheet.cell(row = row, column = time_column).value
                if time_cell == None:
                    continue
                time = self.format_time(str(time_cell))
                if time in times:
                    dates = work_sheet.cell(row = row, column = dates_column).value
                    if self.is_merged(work_sheet, row, dates_column):
                        dates = self.get_value_of_merged_call(work_sheet, row, dates_column)
                    group_name = work_sheet.cell(row = groups_row, column = column).value
                    self.save_subj_in_db(db_name, dates, time, group_name, subject)
                else:
                    print(subject)
        print('ws finished')

    def return_db_name(self, file_name):
        print(213)
        print(file_name)
        if 'zovs_1_kurs' in file_name:
            return 'zovs_1_kurs'
        elif 'zovs_2_kurs' in file_name:
            return 'zovs_2_kurs'
        elif 'zovs_3_kurs' in file_name:
            return 'zovs_3_kurs'
        elif 'zovs_4_kurs' in file_name:
            return 'zovs_4_kurs'
        elif 'lovs_1_kurs' in file_name:
            return 'lovs_1_kurs'
        elif 'lovs_2_kurs' in file_name:
            return 'lovs_2_kurs'
        elif 'lovs_3_kurs' in file_name:
            return 'lovs_3_kurs'
        elif 'lovs_4_kurs' in file_name:
            return 'lovs_4_kurs'
        elif 'imst_1_kurs' in file_name:
            return 'imst_1_kurs'
        elif 'imst_2_kurs' in file_name:
            return 'imst_2_kurs'
        elif 'imst_3_kurs' in file_name:
            return 'imst_3_kurs'
        elif 'imst_4_kurs' in file_name:
            return 'imst_4_kurs'


    def get_dates(self, work_sheet, row, dates_column):
        dates = work_sheet.cell(row = row, column = dates_column).value
        if self.is_merged(work_sheet, row, dates_column):
            dates = self.get_value_of_merged_call(work_sheet, row, dates_column)
        dates = self.format_dates(dates)
        return dates                

    def create_dates_and_times_in_db(self, work_sheet, db_name):
        dates_column = self.const_dates_column
        time_column = self.const_time_column

        first_row = self.find_row_of_first_lesson(work_sheet)
        times = []
        for row in range(1, self.const_quantity_of_rows):
            time_cell = str(work_sheet.cell(row = row, column = time_column).value)
            if time_cell != None and self.is_time(time_cell):
                time_value = self.format_time(time_cell)
                if time_value == '9:45':
                    times = []
                    times.append(time_value)
                    dates = self.get_dates(work_sheet, row, dates_column)
                elif time_value == '11:30' or time_value == '13:30' or time_value == '15:15':
                    times.append(time_value)
                elif time_value == '17:00':
                    times.append(time_value)
                    next_cell = str(work_sheet.cell(row = row + 1, column = time_column).value)
                    next_cell = self.format_time(next_cell) if self.is_time(next_cell) else False

                    after_next_cell = str(work_sheet.cell(row = row + 2, column = time_column).value)
                    after_next_cell = self.format_time(after_next_cell) if self.is_time(after_next_cell) else False

                    if self.is_time(next_cell) or self.is_time(after_next_cell):
                        if next_cell == '9:45' or after_next_cell == '9:45' and next_cell != '18:40':
                            self.save_dates_and_times(db_name, dates, times)
                            continue                               
                        elif next_cell == '18:40' or after_next_cell == '18:40':
                            continue
                    elif next_cell == False and after_next_cell == False:
                        self.save_dates_and_times(db_name, dates, times)
                elif time_value == '18:40':
                    times.append(time_value)
                    self.save_dates_and_times(db_name, dates, times)

    def create_groups_in_db(self, work_book, db_name):
        work_sheet = work_book[work_book.sheetnames[0]]
        list_of_groups = self.return_all_groups_names(work_sheet)
        db_funcs_for_subjects_db.save_groups(db_name, list_of_groups)

    def parse_work_file(self, work_file):
        print(work_file)
        db_name = self.return_db_name(work_file)
        db_funcs_for_subjects_db.drop_db(db_name)
        db_funcs_for_subjects_db.create_db(db_name)

        work_book = load_workbook(work_file)
        print(work_book)
        self.create_groups_in_db(work_book, db_name)

        for ws in work_book.sheetnames:
            work_sheet = work_book[ws]
            ws_date = str(work_sheet)
            month_to_skip = ['09', '10', '11', '12', '01']
            if ws_date[17:19] in month_to_skip:
                print('skipped' + ws_date)
                continue
            self.create_dates_and_times_in_db(work_sheet, db_name)
            self.parse_work_sheet(work_sheet, db_name)

    def parse_work_file_using_name(self, name, route):
        print('Парсер запущен на ' + name)
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        for work_file in work_files:
            if name in work_file:
                db_name = self.return_db_name(work_file)
                work_book = load_workbook(work_file)
                db_funcs_for_subjects_db.drop_db(db_name)
                db_funcs_for_subjects_db.create_db(db_name)
                self.create_groups_in_db(work_book, db_name)
                self.create_dates_and_times_in_db(work_book, db_name)

                for ws in work_book.sheetnames:
                    work_sheet = work_book[ws]
                    self.parse_work_sheet(work_sheet, db_name)

class Excel_parser_undergraduate(Excel_parser):
    # константное значение
    # пока что во всех расписаниях колонка первых групп одинакова 
    const_first_group_column = 4
    const_time_column = 3
    const_dates_column = 1
    const_quantity_of_rows = 50

    def find_number_of_groups_cell_row(self, work_sheet):
        first_group_column = self.const_first_group_column
        for row in range(1, 10):
            viewed_cell = work_sheet.cell(row = row, column = first_group_column).value 
            if type(viewed_cell) == str and 'Группа' in viewed_cell:
                return row

    def return_columns_numbers_of_all_groups_cells(self, work_sheet):
        columns_numbers_of_all_groups_cells = []
        row_number = self.find_number_of_groups_cell_row(work_sheet)
        first_group_column = self.const_first_group_column
        for column in range(first_group_column, 25):
            group_cell = work_sheet.cell(row = row_number, column = column).value
            if type(group_cell) == str and 'Группа' in group_cell:
                columns_numbers_of_all_groups_cells.append(column)
        return columns_numbers_of_all_groups_cells

    def return_all_groups_names(self, work_sheet):
        groups_names = []
        row_number = self.find_number_of_groups_cell_row(work_sheet)
        first_group_column = self.const_first_group_column
        for column in range(first_group_column, 25):
            group_cell = work_sheet.cell(row = row_number, column = column).value
            if type(group_cell) == str and 'Группа' in group_cell:
                group_name = self.format_group_name(group_cell)
                groups_names.append(group_name)
        return groups_names


    def find_row_of_first_lesson(self, work_sheet):
        time_column = self.const_time_column
        times = ['9:45', '09:45', '9.45', '09.45', '9:45:00', '09:45:00', '9.45:00', '09.45:00']
        for row in range(1, 10):
            time_cell = str(work_sheet.cell(row = row, column = time_column).value)
            if time_cell in times:
                return row 

    def get_value_of_merged_call(self, work_sheet, row, column):
        cell = work_sheet.cell(row, column)
        for mergedCell in work_sheet.merged_cells.ranges:
            if (cell.coordinate in mergedCell):
                column_coor = list(mergedCell)[0][1]
                row_coor = list(mergedCell)[1][1]
                result_value = work_sheet.cell(row = row_coor, column = column_coor).value
                return result_value

    def format_group_name(self, group_name):
        if '  ' in group_name:
            group_name = group_name.replace('  ', ' ')
        if ' ' in group_name:
            group_name = group_name.replace(' ', '_')
        if '\n' in group_name:
            group_name = group_name.replace('\n', '')
        group_name = group_name.rstrip()
        return group_name




    def save_subj_in_db(self, db_name, dates, time, group_name, subject):
        dates = self.format_dates(dates)
        time = self.format_time(time)
        group_name = self.format_group_name(group_name)

        for date in dates:
            db_funcs_for_subjects_db.save_subj(db_name, date, time, group_name, subject)

    def parse_work_sheet(self, work_sheet, db_name):
        time_column = self.const_time_column
        dates_column = self.const_dates_column

        groups_column = self.return_columns_numbers_of_all_groups_cells(work_sheet)
        first_row = self.find_row_of_first_lesson(work_sheet)
        groups_row = self.find_number_of_groups_cell_row(work_sheet)

        times = ['9:45', '11:30', '13:30', '15:15', '17:00', '18:40']
        print('ws start')
        for column in groups_column:
            for row in range(first_row, self.const_quantity_of_rows):
                subject = work_sheet.cell(row = row, column = column).value
                if self.is_merged(work_sheet, row, column):
                    subject = self.get_value_of_merged_call(work_sheet, row, column)
                if subject == None:
                    subject = 'нет предмета'
                time_cell = work_sheet.cell(row = row, column = time_column).value
                if time_cell == None:
                    continue
                time = self.format_time(str(work_sheet.cell(row = row, column = time_column).value))
                if time in times:
                    dates = work_sheet.cell(row = row, column = dates_column).value
                    if self.is_merged(work_sheet, row, dates_column):
                        dates = self.get_value_of_merged_call(work_sheet, row, dates_column)
                    group_name = work_sheet.cell(row = groups_row, column = column).value
                    self.save_subj_in_db(db_name, dates, time, group_name, subject)
                else:
                    print(subject)
        print('ws finished')

    def return_db_name(self, file_name):
        if 'zovs_1_kurs' in file_name:
            return 'zovs_1_kurs'
        elif 'zovs_2_kurs' in file_name:
            return 'zovs_2_kurs'
        elif 'zovs_3_kurs' in file_name:
            return 'zovs_3_kurs'
        elif 'zovs_4_kurs' in file_name:
            return 'zovs_4_kurs'
        elif 'lovs_1_kurs' in file_name:
            return 'lovs_1_kurs'
        elif 'lovs_2_kurs' in file_name:
            return 'lovs_2_kurs'
        elif 'lovs_3_kurs' in file_name:
            return 'lovs_3_kurs'
        elif 'lovs_4_kurs' in file_name:
            return 'lovs_4_kurs'


    def get_dates(self, work_sheet, row, dates_column):
        dates = work_sheet.cell(row = row, column = dates_column).value
        if self.is_merged(work_sheet, row, dates_column):
            dates = self.get_value_of_merged_call(work_sheet, row, dates_column)
        dates = self.format_dates(dates)
        return dates                

    def create_dates_and_times_in_db(self, work_book, db_name):
        dates_column = self.const_dates_column
        time_column = self.const_time_column

        for ws in work_book.sheetnames:
            work_sheet = work_book[ws]
            first_row = self.find_row_of_first_lesson(work_sheet)

            times = []
            for row in range(1, self.const_quantity_of_rows):
                time_cell = str(work_sheet.cell(row = row, column = time_column).value)
                if time_cell != None and self.is_time(time_cell):
                    time_value = self.format_time(time_cell)
                    dates = self.get_dates(work_sheet, row, dates_column)
                    if time_value == '9:45':
                        times = []
                        times.append(time_value)
                    elif time_value == '11:30' or time_value == '13:30' or time_value == '15:15':
                        times.append(time_value)
                    elif time_value == '17:00':
                        times.append(time_value)
                        next_cell = str(work_sheet.cell(row = row + 1, column = time_column).value)
                        next_cell = self.format_time(next_cell) if self.is_time(next_cell) else False

                        after_next_cell = str(work_sheet.cell(row = row + 2, column = time_column).value)
                        after_next_cell = self.format_time(after_next_cell) if self.is_time(after_next_cell) else False
                        if self.is_time(next_cell) or self.is_time(after_next_cell):
                            if next_cell == '9:45' or after_next_cell == '9:45':
                                self.save_dates_and_times(db_name, dates, times)
                                continue                               
                            elif next_cell == '18:40' or after_next_cell == '18:40':
                                continue
                        elif next_cell == False and after_next_cell == False:
                            self.save_dates_and_times(db_name, dates, times)
                    elif time_value == '18:40':
                        times.append(time_value)
                        self.save_dates_and_times(db_name, dates, times)

    def create_groups_in_db(self, work_book, db_name):
        work_sheet = work_book[work_book.sheetnames[0]]
        list_of_groups = self.return_all_groups_names(work_sheet)
        db_funcs_for_subjects_db.save_groups(db_name, list_of_groups)

    def parse_work_file(self, work_file):
        db_name = self.return_db_name(work_file)
        db_funcs_for_subjects_db.drop_db(db_name)
        db_funcs_for_subjects_db.create_db(db_name)

        work_book = load_workbook(work_file)
        self.create_groups_in_db(work_book, db_name)
        self.create_dates_and_times_in_db(work_book, db_name)

        for ws in work_book.sheetnames:
            print(ws)
            work_sheet = work_book[ws]
            self.parse_work_sheet(work_sheet, db_name)

    def parse_work_file_using_name(self, name):
        print('Парсер запущен на ' + name)
        work_files = glob.glob('time_tables/full_time_undergraduate/*.xlsx')
        for work_file in work_files:
            if name in work_file:
                db_name = self.return_db_name(work_file)
                work_book = load_workbook(work_file)
                db_funcs_for_subjects_db.drop_db(db_name)
                db_funcs_for_subjects_db.create_db(db_name)
                self.create_groups_in_db(work_book, db_name)
                self.create_dates_and_times_in_db(work_book, db_name)

                for ws in work_book.sheetnames:
                    print(ws)
                    work_sheet = work_book[ws]
                    self.parse_work_sheet(work_sheet, db_name)

    def run_parser(self, route):
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        print(work_files)
        for work_file in work_files:
            print(work_file)
            self.parse_work_file(work_file)

def run_excel_parser_undergraduate():
    parser = Excel_parser_undergraduate()
    parser.run_parser('full_time_undergraduate')
if __name__ == "__main__":
    run_excel_parser_undergraduate()
    #parser = Excel_parser_magistracyasd_fk()
    #parser.run_parser('full_time_magistracy_fk')
    #parser = Excel_parser_magistracy_imist()
    #parser.run_parser('full_time_magistracy_imst')
    #parser = Excel_parser_magistracy_afk()
    #parser.run_parser('full_time_magistracy_afk') 
    #parser = Excel_parser_undergraduate_imst()
    #parser.run_parser('full_time_undergraduate/imst')
    #parser = Excel_parser_undergraduate()
    #parser.run_parser('full_time_undergraduate')

    

#1540 строк кода