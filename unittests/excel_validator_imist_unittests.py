import sys
sys_path = sys.path[0]
path_to_upper_folder = sys_path[:-10]
sys.path.append(path_to_upper_folder)
path_to_excel_validators = path_to_upper_folder+'\excel_validators\\'
sys.path.append(path_to_excel_validators)

import configurations
import excel_validator_imist
import unittest
from openpyxl import load_workbook
import glob


class Test_check_file_name(unittest.TestCase):

    def test_take_correct_name_return_file_name_ok(self):
        obj = excel_validator_imist.Excel_validator_imist()
        result = obj.check_file_name('1_imist')
        self.assertEqual('Имя файла ОК\n', result)

    def test_take_incorrect_name_return_error_with_correct_text(self):
        obj = excel_validator_imist.Excel_validator_imist()
        with self.assertRaises(excel_validator_imist.File_not_valid) as context:
            obj.check_file_name('1_imst')
        self.assertEqual('Ошибка в имени файла 1_imst', str(context.exception))


class Test_check_worksheet_names(unittest.TestCase):

    def test_take_correct_name_return_worksheet_name_ok(self):
        obj = excel_validator_imist.Excel_validator_imist()
        result = obj.check_worksheet_names(
            ['01.02. - 02.03.', '04.05. - 06.07.'])
        self.assertEqual('Имена листов ОК\n', result)

    def test_take_correct_name_with_word_to_skip_return_worksheet_name_ok(self):
        obj = excel_validator_imist.Excel_validator_imist()
        result = obj.check_worksheet_names(
            ['01.02. - 02.03.', '04.05. - 06.07.', 'Ссылки'])
        self.assertEqual('Имена листов ОК\n', result)

    def test_take_incorrect_name_return_error_with_correct_text(self):
        configurations.month_to_skip = ['01']
        obj = excel_validator_imist.Excel_validator_imist()
        with self.assertRaises(excel_validator_imist.File_not_valid) as context:
            obj.check_worksheet_names(['01.02.-02.03.', '04.05. - 06.07.'])
        self.assertEqual('Ошибка в имени листа 01.02.-02.03.\n',
                         str(context.exception))


class Test_check_group_struct(unittest.TestCase):

    def test_take_correct_imist_1_group_name_no_return(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['11.04. - 16.04.']
        obj = excel_validator_imist.Excel_validator_imist()
        result = obj.check_group_struct(
            worksheet, work_file_name, '11.04. - 16.04.')
        self.assertIsNone(result)

    def test_take_incorrect_imist_1_group_name_return_error(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['18.04. - 23.04.']
        obj = excel_validator_imist.Excel_validator_imist()
        with self.assertRaises(excel_validator_imist.File_not_valid) as context:
            obj.check_group_struct(
                worksheet, work_file_name, '18.04. - 23.04.')
        self.assertEqual(
            'Ошибка в структуре группы в D4 в листе 18.04. - 23.04.', str(context.exception))


class Test_check_structure(unittest.TestCase):

    def test_take_correct_imist_1_file_struct_return_correct_text(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_2.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        result = obj.check_structure(work_book, work_file_name)
        self.assertEqual('Структура ОК\n', result)

    def test_take_incorrect_imist_1_file_struct_return_error(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_3.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        with self.assertRaises(excel_validator_imist.File_not_valid) as context:
            obj.check_structure(work_book, work_file_name)
        self.assertEqual(
            'Ошибка в структуре группы в D4 в листе 18.04. - 23.04.', str(context.exception))


class Test_check_cells_with_lessons(unittest.TestCase):

    def test_take_correct_imist_1_cells_with_lessons_return_correct_text(self):
        configurations.month_to_skip = ['01', '02', '03', '05']
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_4.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        result = obj.check_cells_with_lessons(work_book, work_file_name)
        self.assertIsNone(result)

    def test_take_incorrect_imist_1_lesson_cell_with_new_line_return_exception(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_5.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        with self.assertRaises(excel_validator_imist.File_not_valid) as context:
            obj.check_cells_with_lessons(work_book, work_file_name)
        self.assertEqual(
            'Ошибка в ячейке в G7 в листе 11.04. - 16.04. в предмете "[\'Философия\', \'Практическое занятие\', \'\']"', str(context.exception))

    def test_take_incorrect_imist_1_lesson_cell_with_not_existing_subj_return_exception(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_6.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        with self.assertRaises(excel_validator_imist.File_not_valid) as context:
            obj.check_cells_with_lessons(work_book, work_file_name)
        self.assertEqual(
            'Ошибка в ячейке в G8 в листе 11.04. - 16.04. в предмете "Делавые коммуникации"', str(context.exception))

    def test_take_incorrect_imist_1_lesson_cell_with_not_existing_subj_type_return_exception(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_7.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        with self.assertRaises(excel_validator_imist.File_not_valid) as context:
            obj.check_cells_with_lessons(work_book, work_file_name)
        self.assertEqual(
            'Ошибка в ячейке в D7 в листе 11.04. - 16.04. в типе предмете "Практическое занятее"', str(context.exception))

    def test_take_incorrect_imist_1_location_cell_with_not_existing_location_return_exception(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_8.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        with self.assertRaises(excel_validator_imist.File_not_valid) as context:
            obj.check_cells_with_lessons(work_book, work_file_name)
        self.assertEqual(
            'Ошибка в ячейке в E18 в листе 11.04. - 16.04. в локации "ауд. 500"', str(context.exception))

    def test_take_incorrect_imist_1_teacher_cell_with_not_existing_teacher_return_exception(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_9.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        with self.assertRaises(excel_validator_imist.File_not_valid) as context:
            obj.check_cells_with_lessons(work_book, work_file_name)
        self.assertEqual(
            'Ошибка в ячейке в F6 в листе 11.04. - 16.04. в преподавателе [\'ВВВВВВВ\']', str(context.exception))

    def test_take_incorrect_imist_1_second_teacher_cell_with_not_existing_teacher_return_exception(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_10.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        with self.assertRaises(excel_validator_imist.File_not_valid) as context:
            obj.check_cells_with_lessons(work_book, work_file_name)
        self.assertEqual(
            'Ошибка в ячейке в I11 в листе 11.04. - 16.04. в преподавателе [\'Ермилова В.В.\', \'В.А.\']', str(context.exception))

    def test_take_timtable_with_holiday_return_no_exception(self):
        configurations.month_to_skip = ['01', '02', '03', '05']
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_13.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        result = obj.check_cells_with_lessons(work_book, work_file_name)
        self.assertIsNone(result)

    def test_take_timtable_with_event_return_no_exception(self):
        configurations.month_to_skip = ['01', '02', '03', '05']
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_14.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        result = obj.check_cells_with_lessons(work_book, work_file_name)
        self.assertIsNone(result)


class Test_check_practice_cell(unittest.TestCase):

    def test_take_correct_imist_1_practice_no_return(self):
        configurations.month_to_skip = ['01', '02', '03', '05']
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_11.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        result = obj.check_cells_with_lessons(work_book, work_file_name)
        self.assertIsNone(result)

    def test_take_incorrect_imist_1_practice_returrn_exception(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate_imist/1_imist_12.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_imist.Excel_validator_imist()
        with self.assertRaises(excel_validator_imist.File_not_valid) as context:
            obj.check_cells_with_lessons(work_book, work_file_name)
        self.assertEqual(
            'Ошибка в практике в D6 в листе 11.04. - 16.04. в датах "21.03. - 03.05"', str(context.exception))


if __name__ == '__main__':
    unittest.main()
