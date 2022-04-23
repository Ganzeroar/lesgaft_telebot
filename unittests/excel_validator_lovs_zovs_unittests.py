import unittest
from openpyxl import Workbook, load_workbook, utils
import glob
from unittest.mock import patch

import sys
sys_path = sys.path[0]
path_to_upper_folder = sys_path[:-10]
sys.path.append(path_to_upper_folder)
path_to_excel_validators = path_to_upper_folder+'\excel_validators\\'
sys.path.append(path_to_excel_validators)

import excel_validator_lovs_zovs
import configurations

class Test_check_file_name(unittest.TestCase):

    def test_take_correct_name_return_file_name_ok(self):
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        result = obj.check_file_name('1_lovs')
        self.assertEqual('Имя файла ОК\n', result)

    def test_take_incorrect_name_return_error_with_correct_text(self):
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        with self.assertRaises(excel_validator_lovs_zovs.File_not_valid) as context:
            obj.check_file_name('1_ lovs')
        self.assertEqual('Ошибка в имени файла 1_ lovs', str(context.exception))

class Test_check_worksheet_names(unittest.TestCase):

    def test_take_correct_name_return_worksheet_name_ok(self):
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        result = obj.check_worksheet_names(['01.02. - 02.03.', '04.05. - 06.07.'])
        self.assertEqual('Имена листов ОК\n', result)

    def test_take_correct_name_with_word_to_skip_return_worksheet_name_ok(self):
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        result = obj.check_worksheet_names(['01.02. - 02.03.', '04.05. - 06.07.', 'Ссылки'])
        self.assertEqual('Имена листов ОК\n', result)

    def test_take_incorrect_name_return_error_with_correct_text(self):
        configurations.month_to_skip = ['01']
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        with self.assertRaises(excel_validator_lovs_zovs.File_not_valid) as context:
            obj.check_worksheet_names(['01.02.-02.03.', '04.05. - 06.07.'])
        self.assertEqual('Ошибка в имени листа 01.02.-02.03.\n', str(context.exception))

class Test_check_group_struct(unittest.TestCase):

    def test_take_correct_lovs_1_group_name_no_return(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['11.04. - 16.04.']
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        result = obj.check_group_struct(worksheet, work_file_name, '11.04. - 16.04.')
        self.assertIsNone(result)
    
    def test_take_correct_lovs_3_group_name_no_return(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/3_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['11.04. - 16.04.']
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        result = obj.check_group_struct(worksheet, work_file_name, '11.04. - 16.04.')
        self.assertIsNone(result)

    def test_take_incorrect_lovs_1_group_name_return_error(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['18.04. - 23.04.']
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        with self.assertRaises(excel_validator_lovs_zovs.File_not_valid) as context:
            obj.check_group_struct(worksheet, work_file_name, '18.04. - 23.04.')
        self.assertEqual('Ошибка в структуре группы в D4 в листе 18.04. - 23.04.', str(context.exception))
    
    def test_take_correct_lovs_1_group_specialication_no_return(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['11.04. - 16.04.']
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        result = obj.check_group_struct(worksheet, work_file_name, '11.04. - 16.04.')
        self.assertIsNone(result)

    def test_take_incorrect_lovs_1_group_specialication_return_error(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['25.04. - 30.04.']
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        with self.assertRaises(excel_validator_lovs_zovs.File_not_valid) as context:
            obj.check_group_struct(worksheet, work_file_name, '25.04. - 30.04.')
        self.assertEqual('Ошибка в специализации в D5 в листе 25.04. - 30.04.', str(context.exception))

class Test_check_date_struct(unittest.TestCase):

    def test_take_correct_lovs_1_date_struct_no_return(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['11.04. - 16.04.']
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        result = obj.check_date_struct(worksheet, work_file_name, '11.04. - 16.04.')
        self.assertIsNone(result)

    def test_take_incorrect_lovs_1_date_struct_return_error(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['25.04. - 30.04.']
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        with self.assertRaises(excel_validator_lovs_zovs.File_not_valid) as context:
            obj.check_date_struct(worksheet, work_file_name, '25.04. - 30.04.')
        self.assertEqual('Ошибка в структуре даты в A26 в листе 25.04. - 30.04.', str(context.exception))

class Test_check_day_struct(unittest.TestCase):

    def test_take_correct_lovs_1_day_struct_no_return(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['11.04. - 16.04.']
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        result = obj.check_day_struct(worksheet, work_file_name, '11.04. - 16.04.')
        self.assertIsNone(result)

    def test_take_incorrect_lovs_1_day_struct_return_error(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['18.04. - 23.04.']
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        with self.assertRaises(excel_validator_lovs_zovs.File_not_valid) as context:
            obj.check_day_struct(worksheet, work_file_name, '18.04. - 23.04.')
        self.assertEqual('Ошибка в структуре дня в B11 в листе 18.04. - 23.04.', str(context.exception))

class Test_check_time_struct(unittest.TestCase):

    def test_take_correct_lovs_1_time_struct_no_return(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['11.04. - 16.04.']
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        result = obj.check_time_struct(worksheet, work_file_name, '11.04. - 16.04.')
        self.assertIsNone(result)

    def test_take_incorrect_lovs_1_time_struct_return_error(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['18.04. - 23.04.']
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        with self.assertRaises(excel_validator_lovs_zovs.File_not_valid) as context:
            obj.check_time_struct(worksheet, work_file_name, '18.04. - 23.04.')
        self.assertEqual('Ошибка в структуре времени в C14 в листе 18.04. - 23.04.', str(context.exception))

class Test_check_structure(unittest.TestCase):

    def test_take_correct_lovs_1_file_struct_return_correct_text(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/1_lovs_2.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        result = obj.check_structure(work_book, work_file_name)
        self.assertEqual('Структура ОК\n', result)

    def test_take_incorrect_lovs_1_file_struct_return_error(self):
        work_file = glob.glob(f'test_time_tables/full_time_undergraduate/1_lovs_3.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        obj = excel_validator_lovs_zovs.Excel_validator_lovs_zovs()
        with self.assertRaises(excel_validator_lovs_zovs.File_not_valid) as context:
            obj.check_structure(work_book, work_file_name)
        self.assertEqual('Ошибка в специализации в D5 в листе 25.04. - 30.04.', str(context.exception))




if __name__ == '__main__':
    unittest.main()