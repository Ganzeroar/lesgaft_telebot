import sys
sys_path = sys.path[0]
path_to_upper_folder = sys_path[:-10]
sys.path.append(path_to_upper_folder)
path_to_excel_validators = path_to_upper_folder+'\excel_validators\\'
sys.path.append(path_to_excel_validators)

import configurations
import excel_validator
import unittest
from openpyxl import load_workbook
import glob
from file_not_valid_exception import File_not_valid


class Test_excel_validator(unittest.TestCase):
    
    @unittest.skip("passed")
    def test_take_one_correct_and_one_incorrect_sheetname_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        result = obj.run_validator('test_time_tables/full_time_undergraduate')
        self.assertEqual('Имя файла ОК\nИмена листов ОК\nЛист 21.03. - 26.03.\nСтруктура групп ОК\nСруктура дат ОК\nСруктура дней ОК\nСруктура времени ОК\nЛист 21.03. - 26.03.\nИмена групп ОК\nДаты ОК\nИмена дней ОК\nВремя ОК\n', result)

    # def test_take_full_file_return_correct_message(self):
    #    obj = excel_validator.Excel_validator()
    #    result = obj.run_validator('test_time_tables/full_time_undergraduate/fulltest')
    #    self.assertEqual('ОК\n', result)


class Test_check_worksheet_name(unittest.TestCase):
    def test_take_correct_sheetnames_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        result = obj.check_worksheet_names(
            ['24.01. - 29.01.', '31.01. - 05.02.'])
        self.assertEqual('Имена листов ОК\n', result)

    def test_take_one_incorrect_sheetname_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        with self.assertRaises(File_not_valid) as context:
            obj.check_worksheet_names(['24.01 - 29.01'])
        self.assertEqual(
            'Ошибка в имени листа 24.01 - 29.01\n', str(context.exception))

    def test_take_all_incorrect_sheetname_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        with self.assertRaises(File_not_valid) as context:
            obj.check_worksheet_names(['24.01 - 29.01', '31.01 - 05.02'])
        self.assertEqual(
            'Ошибка в имени листа 24.01 - 29.01\n', str(context.exception))


class Test_check_date_struct(unittest.TestCase):
    def test_take_correct_date_struct_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate/1_lovs_6.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book[work_book.sheetnames[0]]

        constants = configurations.group_constants['lovs_1']
        result = obj.check_date_struct(worksheet, work_book.sheetnames[0], constants)
        self.assertIsNone(result)

    def test_take_incorrect_lovs_1_date_struct_return_error(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['25.04. - 30.04.']
        obj = excel_validator.Excel_validator()
        constants = configurations.group_constants['lovs_1']
        with self.assertRaises(File_not_valid) as context:
            obj.check_date_struct(worksheet, '25.04. - 30.04.', constants)
        self.assertEqual(
            'Ошибка в структуре даты в A26 в листе 25.04. - 30.04.', str(context.exception))



class Test_check_day_struct(unittest.TestCase):
    def test_take_correct_day_struct_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate/1_lovs_5.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book[work_book.sheetnames[0]]

        constants = configurations.group_constants['lovs_1']
        result = obj.check_day_struct(worksheet, work_book.sheetnames[0], constants)
        self.assertIsNone(result)

    def test_take_incorrect_lovs_1_day_struct_return_error(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['18.04. - 23.04.']
        obj = excel_validator.Excel_validator()
        constants = configurations.group_constants['lovs_1']
        with self.assertRaises(File_not_valid) as context:
            obj.check_day_struct(worksheet, '18.04. - 23.04.', constants)
        self.assertEqual(
            'Ошибка в структуре дня в B11 в листе 18.04. - 23.04.', str(context.exception))

class Test_check_time_struct(unittest.TestCase):
    def test_take_correct_time_struct_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate/1_lovs_4.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book[work_book.sheetnames[0]]

        constants = configurations.group_constants['lovs_1']
        result = obj.check_time_struct(worksheet, work_book.sheetnames[0], constants)
        self.assertIsNone(result)

    def test_take_incorrect_lovs_1_time_struct_return_error(self):
        work_file = glob.glob(
            f'test_time_tables/full_time_undergraduate/1_lovs_1.xlsx')
        work_file_name = work_file[0]
        work_book = load_workbook(work_file_name)
        worksheet = work_book['18.04. - 23.04.']
        obj = excel_validator.Excel_validator()
        constants = configurations.group_constants['lovs_1']
        with self.assertRaises(File_not_valid) as context:
            obj.check_time_struct(worksheet, '18.04. - 23.04.', constants)
        self.assertEqual(
            'Ошибка в структуре времени в C14 в листе 18.04. - 23.04.', str(context.exception))



if __name__ == '__main__':
    unittest.main()
