
from openpyxl import Workbook, load_workbook, utils
import glob
import re
import configurations
import os

class Excel_validator():
    const_group_column = 4
    const_group_row = 4

    #TODO Если нужна такая фича, могу запилить что-то типа кнопки "свободные аудитории", по которой бот отвечает что сейчас никем по расписанию не занято

    def run_validator_for_excel_parser(self, route):
        result_message = ''
        #перебирать листы
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        #если в имени файла ошибка - выбрасывать
        for work_file in work_files:
            result_message += f'{work_file}\n'
            result_message += self.check_file_name(work_file)
            if 'Ошибка' in result_message:
                return result_message
            work_book = load_workbook(work_file)
            result_message += self.check_worksheet_names(work_book.sheetnames)
            if 'Ошибка' in result_message:
                return result_message
            #если неверная структура - выбрасывать
            result_message += self.check_structure(work_book, work_file)
            if 'Ошибка' in result_message:
                return result_message
            result_message += self.check_content_of_servise_cells(work_book, work_file)
            if 'Ошибка' in result_message:
                return result_message
            #result_message += self.check_class_schedule(work_book, work_file)
            print(f'{work_file} валиден')
        
        return result_message


    def run_validator(self, route):
        result_message = ''
        #перебирать листы
        file_name = glob.glob(f'time_tables/{route}/*.xlsx')[0]
        #если в имени файла ошибка - выбрасывать
        result_message += self.check_file_name(file_name)
        if 'Ошибка' in result_message:
            path = os.path.join(os.path.abspath(os.path.dirname(__file__)), file_name)
            os.remove(path)
            return result_message
        work_book = load_workbook(file_name)
        result_message += self.check_worksheet_names(work_book.sheetnames)
        if 'Ошибка' in result_message:
            path = os.path.join(os.path.abspath(os.path.dirname(__file__)), file_name)
            os.remove(path)
            return result_message
        #если неверная структура - выбрасывать
        result_message += self.check_structure(work_book, file_name)
        if 'Ошибка' in result_message:
            path = os.path.join(os.path.abspath(os.path.dirname(__file__)), file_name)
            os.remove(path)
            return result_message
        result_message += self.check_content_of_servise_cells(work_book, file_name)
        #result_message += self.check_class_schedule(work_book, file_name)
        path = os.path.join(os.path.abspath(os.path.dirname(__file__)), file_name)
        os.remove(path)
        print(result_message)
        return result_message

    def check_class_schedule(self, work_book, file_name):
        message = ''
        for worksheet_name in work_book.sheetnames:
            if worksheet_name.lower() in configurations.words_to_skip:
                continue
            if self.is_reason_to_skip(worksheet_name.lower()) == True:
                continue
            message += f'Лист {worksheet_name}\n'
            worksheet = work_book[worksheet_name]
            try:
                message += self.check_worksheet_schedule(worksheet, file_name)
            except:
                message += 'Критическая Ошибка'
            if 'Ошибка' in message:
                return message

        return message

    def is_reason_to_skip(self, worksheet_name):
        month_to_skip = configurations.month_to_skip
        month = worksheet_name[-3:-1]
        #day = worksheet_name[9:11]
        if month in month_to_skip:
        #    if day in configurations.day_to_parse:
        #        return False
            return True
        return False
        

    def check_worksheet_schedule(self, worksheet, file_name):
        message = ''
        constants = self.return_current_file_constants(file_name)
        const_first_class_cell_column = constants['first_group_column']        
        const_last_class_cell_column = constants['last_group_column']
        const_first_class_cell_row = constants['first_time_row']
        error_values = []
        for row in range(const_first_class_cell_row, 50):
            for column in range(const_first_class_cell_column, const_last_class_cell_column + 1):
                viewed_class_cell = worksheet.cell(row = row, column = column)
                if self.is_merged(worksheet, viewed_class_cell):
                    viewed_class_value = self.get_merged_cell_value(worksheet, viewed_class_cell)
                else:
                    viewed_class_value = viewed_class_cell.value
                if viewed_class_value == None:
                    continue
                cell_data = viewed_class_value.split('\n')
                if '' in cell_data or ' ' in cell_data:
                    message += f'Пустая строка в {viewed_class_cell.coordinate}\n'
                    continue
                if '  ' in cell_data[0]:
                    message += f'Минимум 2 пробела в {viewed_class_cell.coordinate}\n'
                    continue
                if cell_data[0] == 'Начальник УМУ Паульс А.А.':
                    if message == '':
                        message += 'Предметы ОК\n'
                    return message
                #print(viewed_class_cell.coordinate)
                #print(viewed_class_value)
                if 'ТиМ ИВС' in cell_data:
                    continue
                    #TODO а что если перед или после тире не будет пробела?
                    if len(cell_data) == 3:
                        subject = cell_data[0]
                        first_teacher = cell_data[1].split(' - ')[1]
                        second_teacher = cell_data[2].split(' - ')[1]
                        if subject not in configurations.existing_subjects:
                            message += f'Ошибка в предмете в  {viewed_class_cell.coordinate}\n'
                        if first_teacher not in configurations.existing_teachers:
                            message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                        if second_teacher not in configurations.existing_teachers:
                            message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                elif 'Элективные дисциплины по ФКиС' in cell_data:
                    continue
                    if len(cell_data) == 3:
                        subject = cell_data[0]
                        first_teacher = cell_data[1].split(' - ')[1]
                        second_teacher = cell_data[2].split(' - ')[1]
                        if subject not in configurations.existing_subjects:
                            message += f'Ошибка в предмете в  {viewed_class_cell.coordinate}\n'
                        if first_teacher not in configurations.existing_teachers:
                            message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                        if second_teacher not in configurations.existing_teachers:
                            message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                elif cell_data[0] in configurations.strings_to_skip_while_no_format:
                    continue
                elif len(cell_data) == 1:
                    continue
                elif len(cell_data) > 1:
                    if cell_data[1] in configurations.strings_to_skip_while_no_format:
                        continue
                elif 'практика' in cell_data[1]:
                    dates = cell_data[0]
                    practice = cell_data[1]
                    result = re.fullmatch(r'С\s\d{2}[.]\d{2}[.]\sпо\s\d{2}[.]\d{2}[.]', dates)
                    if result == None:
                        message += f'Ошибка в датах в {viewed_class_cell.coordinate}\n'
                    if practice not in configurations.existing_practice:
                        message += f'Ошибка в практике в {viewed_class_cell.coordinate}\n'
                elif len(cell_data) == 4:
                    location = cell_data[0]
                    subject = cell_data[1]
                    type_of_subject = cell_data[2]
                    teacher = cell_data[3]
                    if 'ауд' in location:
                        result = re.fullmatch(r'ауд.\s\d{1,3}', location)
                        if result == None:
                            result_2 = re.fullmatch(r'ауд.\s\d{1,3}\sИМиСТ', location)
                            if result_2 == None:
                                message += f'Ошибка в формате аудитории в  {viewed_class_cell.coordinate}\n'
                        auditorium = location[5:]
                        if auditorium not in configurations.existing_locations:
                            message += f'Ошибка в аудитории в  {viewed_class_cell.coordinate}\n'
                    if subject not in configurations.existing_subjects:
                        message += f'Ошибка в предмете в  {viewed_class_cell.coordinate}\n'
                    if type_of_subject not in configurations.existing_type_of_subject_type_1:
                        message += f'Ошибка в типе предмета в  {viewed_class_cell.coordinate}\n'
                    if teacher not in configurations.existing_teachers:
                        message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                elif len(cell_data) == 3:
                    # ловс 2 тут вопрос не решён
                    location = cell_data[0]
                    subject = cell_data[1]
                    teacher = cell_data[2]
                    if location not in configurations.existing_locations:
                        message += f'Ошибка в локации в  {viewed_class_cell.coordinate}\n'
                    if subject not in configurations.existing_subjects:
                        message += f'Ошибка в предмете в  {viewed_class_cell.coordinate}\n'
                    if teacher not in configurations.existing_teachers:
                        message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                elif len(cell_data) == 2:
                    subject = cell_data[0]
                    teacher = cell_data[1]
                    if subject in configurations.strings_to_skip_while_no_format:
                        continue
                    #if ' - ' in teacher:
                        
                    if subject not in configurations.existing_subjects:
                        message += f'Ошибка в предмете в  {viewed_class_cell.coordinate}\n'
                    if teacher not in configurations.existing_teachers:
                        message += f'Ошибка в преподавателе в  {viewed_class_cell.coordinate}\n'
                if len(message) >= 2000:
                    return message
        if message == '':
            message += 'Предметы ОК'
        return message



        return message

    def check_content_of_servise_cells(self, work_book, file_name):
        message = ''
        for worksheet_name in work_book.sheetnames:
            if worksheet_name.lower() in configurations.words_to_skip:
                continue
            if self.is_reason_to_skip(worksheet_name.lower()) == True:
                continue
            worksheet = work_book[worksheet_name]
            status = self.check_group_numbers(worksheet, file_name)
            if 'Ошибка' in status:
                message += f'Лист {worksheet_name}\n' + status
                return message
            status = self.check_group_specializations(worksheet, file_name)
            if 'Ошибка' in status:
                message += f'Лист {worksheet_name}\n' + status
                return message
            status = self.check_date_column(worksheet, file_name)
            if 'Ошибка' in status:
                message += f'Лист {worksheet_name}\n' + status
                return message
            status = self.check_day_of_week_column(worksheet, file_name)
            if 'Ошибка' in status:
                message += f'Лист {worksheet_name}\n' + status
                return message
            status = self.check_time_column(worksheet, file_name)
            if 'Ошибка' in status:
                return message
        if message == '':
            message += 'Служебные ячейки ОК\n'
        return message

    def check_time_column(self, worksheet, file_name):
        message = ''
        constants = self.return_current_file_constants(file_name)
        const_time_column = constants['time_column']
        const_first_time_row = constants['first_time_row']
        lesson_start_times = configurations.lesson_start_times
        for row in range(const_first_time_row, 50):
            viewed_time_row =  row
            viewed_time_cell = worksheet.cell(row = viewed_time_row, column = const_time_column)
            viewed_time_value = viewed_time_cell.value
            if self.is_merged(worksheet, viewed_time_cell) == True:
                if self.get_merged_cell_value(worksheet, viewed_time_cell) == 'Начальник УМУ Паульс А.А.':
                    break
            if viewed_time_value in lesson_start_times:
                continue
            else:
                message += f'Ошибка в времени в {viewed_time_cell.coordinate}\n'
                return message
        return message

    def check_day_of_week_column(self, worksheet, file_name):
        message = ''
        constants = self.return_current_file_constants(file_name)
        const_day_column = constants['day_column']
        const_first_day_row = constants['first_day_row']
        days_of_week = configurations.days_of_week
        for row in range(const_first_day_row, 50):
            viewed_day_row =  row
            viewed_day_cell = worksheet.cell(row = viewed_day_row, column = const_day_column)
            viewed_day_value = self.get_merged_cell_value(worksheet, viewed_day_cell)
            if viewed_day_value == 'Начальник УМУ Паульс А.А.':
                break
            if viewed_day_value in days_of_week:
                continue
            else:
                message += f'Ошибка в дне в {viewed_day_cell.coordinate}\n'
                return message
        return message
        

    def check_date_column(self, worksheet, file_name):
        message = ''
        constants = self.return_current_file_constants(file_name)
        const_date_column = constants['date_column']
        const_first_date_row = constants['first_date_row']
        current_date = worksheet.cell(row = const_first_date_row, column = const_date_column).value
        current_day = int(current_date[:2])
        current_month = int(current_date[-3:-1])

        for row in range(const_first_date_row, 50):
            viewed_date_row = row
            viewed_date_cell = worksheet.cell(row = viewed_date_row, column = const_date_column)
            viewed_date_value = self.get_merged_cell_value(worksheet, viewed_date_cell)
            if viewed_date_value in configurations.string_for_stop_vertical_parsing:
                break
            viewed_day = int(viewed_date_value[:2])
            viewed_month = int(viewed_date_value[-3:-1])
            if viewed_day == current_day and current_month == viewed_month:
                continue
            elif viewed_day == current_day + 1 and current_month == viewed_month:
                current_day += 1
                continue
            elif viewed_month == current_month + 1:
                current_month += 1
                current_day = 1
            else:
                message += f'Ошибка в дате в {viewed_date_cell.coordinate}\n'
                return message
        return message

    def check_group_numbers(self, worksheet, file_name):
        message = ''
        constants = self.return_current_file_constants(file_name)
        const_group_row = constants['group_row']
        const_first_group_column = constants['first_group_column']
        const_group_numbers = constants['group_numbers']
        const_last_group_column = constants['last_group_column']
        group_range = const_last_group_column - const_first_group_column
        # проверить последний столбик
        for i in range(0, group_range + 1):
            group_column = const_first_group_column + i
            viewed_group_cell = worksheet.cell(row = const_group_row, column = group_column)
            if self.is_merged(worksheet, viewed_group_cell):
                viewed_group_value = self.get_merged_cell_value(worksheet, viewed_group_cell)
            else:
                viewed_group_value = viewed_group_cell.value
            if const_group_numbers[i] != viewed_group_value:
                message += f'Ошибка в имени группы в {viewed_group_cell.coordinate}\n'
        return message

    def check_group_specializations(self, worksheet, file_name):
        message = ''
        constants = self.return_current_file_constants(file_name)
        const_group_specialization_row = constants['group_specialization_row']
        const_group_specializations = constants['specialization']
        const_first_group_column = constants['first_group_column']
        const_last_group_column = constants['last_group_column']
        group_range = const_last_group_column - const_first_group_column
        for i in range(0, group_range + 1):
            group_column = const_first_group_column + i
            viewed_group_cell = worksheet.cell(row = const_group_specialization_row, column = group_column)
            if const_group_specializations[i] != viewed_group_cell.value:
                message += f'Ошибка в специализации в {viewed_group_cell.coordinate}\n'
        return message

    def check_structure(self, work_book, file_name):
        message = ''
        for worksheet_name in work_book.sheetnames:
            if worksheet_name.lower() in configurations.words_to_skip:
                continue
            if self.is_reason_to_skip(worksheet_name.lower()) == True:
                continue
            worksheet = work_book[worksheet_name]
            status = self.check_group_struct(worksheet, file_name)
            if 'Ошибка' in status:
                message += f'Лист {worksheet_name}\n' + status
                return message
            status = self.check_date_struct(worksheet, file_name)
            if 'Ошибка' in status:
                message += f'Лист {worksheet_name}\n' + status
                return message
            status = self.check_day_struct(worksheet, file_name)
            if 'Ошибка' in status:
                message += f'Лист {worksheet_name}\n' + status
                return message
            status = self.check_time_struct(worksheet, file_name)
            if 'Ошибка' in status:
                message += f'Лист {worksheet_name}\n' + status
                return message
        if message == '':
            message = 'Структура ОК\n'
        return message

    def check_time_struct(self, worksheet, file_name):
        message = ''
        constants = self.return_current_file_constants(file_name)
        const_time_column = constants['time_column']
        const_first_time_row = constants['first_time_row']
        
        first_time = worksheet.cell(row = const_first_time_row, column = const_time_column)
        if isinstance(first_time.value, str) == False:
            message += f'Ошибка неверный тип данных в {first_time.coordinate}\n'
            return message
        result = re.fullmatch(r'\d{1,2}[:]\d{2}', first_time.value)
        if result == None:
            message += f'Ошибка в первом времени в {first_time.coordinate}\n'
        for row in range(0, 50):
            viewed_time_row = const_first_time_row + row
            viewed_time_cell = worksheet.cell(row = viewed_time_row, column = const_time_column)
            if self.is_merged(worksheet, viewed_time_cell):
                viewed_time_value = self.get_merged_cell_value(worksheet, viewed_time_cell)
                if viewed_time_value in configurations.string_for_stop_vertical_parsing:
                    break
                else:
                    message += f'Ошибка пустая строка в {viewed_time_cell.coordinate}\n'
                    return message
            else:
                viewed_time_value = viewed_time_cell.value
            if viewed_time_value == None:
                message += f'Ошибка пустая строка в {viewed_time_cell.coordinate}\n'
                return message
            if isinstance(viewed_time_value, str) == False:
                message += f'Ошибка неверный тип данных в {viewed_time_cell.coordinate}\n'
                return message
            result = re.fullmatch(r'\d{1,2}[:]\d{2}', viewed_time_value)
            if result == None:
                message += f'Ошибка во времени в {viewed_time_cell.coordinate}\n'

        for row in range(0, 50):
            viewed_time_row = const_first_time_row + row
            viewed_time_cell = worksheet.cell(row = viewed_time_row, column = const_time_column)
            if self.is_merged(worksheet, viewed_time_cell):
                viewed_time_value = self.get_merged_cell_value(worksheet, viewed_time_cell)
                if viewed_time_value in configurations.string_for_stop_vertical_parsing:
                    break
                else:
                    message += f'Ошибка пустая строка в {viewed_time_cell.coordinate}\n'
                    return message
            else:
                viewed_time_value = viewed_time_cell.value

            next_viewed_time_cell = worksheet.cell(row = viewed_time_row+1, column = const_time_column)
            if self.is_merged(worksheet, next_viewed_time_cell):
                next_viewed_time_value = self.get_merged_cell_value(worksheet, next_viewed_time_cell)
                if next_viewed_time_value in configurations.string_for_stop_vertical_parsing:
                    break
                else:
                    message += f'Ошибка пустая строка в {next_viewed_time_cell.coordinate}\n'
                    return message
            else:
                next_viewed_time_value = next_viewed_time_cell.value
            if viewed_time_value == '9:45':
                if next_viewed_time_value == '11:30':
                    continue
                else:
                    message += f'Ошибка неверная следующая ячейка в {viewed_time_cell.coordinate}\n'
                    return message
            elif viewed_time_value == '11:30':
                if next_viewed_time_value == '13:30':
                    continue
                else:
                    message += f'Ошибка неверная следующая ячейка в {viewed_time_cell.coordinate}\n'
                    return message
            elif viewed_time_value == '13:30':
                if next_viewed_time_value == '15:15':
                    continue
                else:
                    message += f'Ошибка неверная следующая ячейка в {viewed_time_cell.coordinate}\n'
                    return message
            elif viewed_time_value == '15:15':
                if next_viewed_time_value == '17:00':
                    continue
                else:
                    message += f'Ошибка неверная следующая ячейка в {viewed_time_cell.coordinate}\n'
                    return message
            elif viewed_time_value == '17:00':
                if next_viewed_time_value == '9:45':
                    continue
                elif next_viewed_time_value == '18:40':
                    continue
                else:
                    message += f'Ошибка неверная следующая ячейка в {viewed_time_cell.coordinate}\n'
                    return message
            elif viewed_time_value == '18:40':
                if next_viewed_time_value == '9:45':
                    continue
                else:
                    message += f'Ошибка неверная следующая ячейка в {viewed_time_cell.coordinate}\n'
                    return message

        
        return message

    def get_value_of_cell(self, worksheet, row, column):
        #пока не используется
        viewed_cell = worksheet.cell(row = row, column = column)
        if self.is_merged(worksheet, viewed_cell):
            viewed_cell_value = self.get_merged_cell_value(worksheet, viewed_cell)
            #if viewed_time_value in configurations.string_for_stop_vertical_parsing:
            #    break
        else:
            viewed_cell_value = viewed_cell.value
        return viewed_cell_value
        

    def check_day_struct(self, worksheet, file_name):
        message = ''
        constants = self.return_current_file_constants(file_name)
        const_day_column = constants['day_column']
        const_first_day_row = constants['first_day_row']
        
        for row in range(0, 50):
            viewed_day_row = const_first_day_row + row
            viewed_day_cell = worksheet.cell(row = viewed_day_row, column = const_day_column)
            if self.is_merged(worksheet, viewed_day_cell):
                viewed_day_value = self.get_merged_cell_value(worksheet, viewed_day_cell)
            else:
                viewed_day_value = viewed_day_cell.value
            if viewed_day_value in configurations.string_for_stop_vertical_parsing:
                break
                #message += f'Ошибка в последней строке в {viewed_day_cell.coordinate}'
                #return message
            if viewed_day_value == None:
                message += f'Ошибка пустая строка в {viewed_day_cell.coordinate}'
                return message
            result = re.fullmatch(r'[А-Я][а-я][.]', viewed_day_value)
            if result == None:
                message += f'Ошибка в дне в {viewed_day_cell.coordinate}\n'
        return message
        

    def check_date_struct(self, worksheet, file_name):
        message = ''
        constants = self.return_current_file_constants(file_name)
        const_date_column = constants['date_column']
        const_first_date_row = constants['first_date_row']
        for row in range(0, 50):
            viewed_date_row = const_first_date_row + row
            viewed_date_cell = worksheet.cell(row = viewed_date_row, column = const_date_column)
            if self.is_merged(worksheet, viewed_date_cell):
                viewed_date_value = self.get_merged_cell_value(worksheet, viewed_date_cell)
            else:
                viewed_date_value = viewed_date_cell.value
            if viewed_date_value == None:
                message += f'Пустая строка в {viewed_date_cell.coordinate}\n'
                return message
            if viewed_date_value in configurations.string_for_stop_vertical_parsing:
                break
            result = re.fullmatch(r'\d{2}[.]\d{2}[.]', viewed_date_value)
            if result == None:
                if 'Паульс' in viewed_date_value:
                    message += f'Ошибка в последней строке в {viewed_date_cell.coordinate}\n'
                message += f'Ошибка в структуре даты в {viewed_date_cell.coordinate}\n'
        return message
        
    def is_merged(self, worksheet, viewed_cell):
        for mergedCell in worksheet.merged_cells.ranges:
            if (viewed_cell.coordinate in mergedCell):
                return True
        return False

    def get_merged_cell_value(self, worksheet, cell):
        rng = []
        for s in worksheet.merged_cells.ranges:
            if cell.coordinate in s:
                rng.append(s)
        cell_value = worksheet.cell(rng[0].min_row, rng[0].min_col).value
        return cell_value
        #rng = [s for s in worksheet.merged_cells.ranges if cell.coordinate in s]
        #return worksheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng)!=0 else cell.value

    def check_group_struct(self, worksheet, file_name):
        message = ''
        result = ''
        constants = self.return_current_file_constants(file_name)
        const_group_row = constants['group_row']
        const_first_group_column = constants['first_group_column']
        const_last_group_column = constants['last_group_column']
        for column in range(const_first_group_column, const_last_group_column + 1):
            viewed_group_cell = worksheet.cell(row = const_group_row, column = column)
            viewed_group_value = viewed_group_cell.value
            if viewed_group_value == None:
                if self.is_merged(worksheet, viewed_group_cell) == True:
                    try:
                        const_merged_group = constants['merged_group']
                        viewed_group_value = self.get_merged_cell_value(worksheet, viewed_group_cell)
                        if viewed_group_value != const_merged_group:
                            message += f'Объединённая ячейка в {viewed_group_cell.coordinate}\n'
                    except:
                        message += f'Объединённая ячейка в {viewed_group_cell.coordinate}\n'
                else:
                    message += f'Пустая строка в {viewed_group_cell.coordinate}\n'
                return message
            try:
                result = re.fullmatch(r'Группа\s\d{3}', viewed_group_value)
            except:
                message += f'Ошибка в структуре группы в {viewed_group_cell.coordinate}\n'
            if result == None:
                message += f'Ошибка в структуре группы в {viewed_group_cell.coordinate}\n'
        return message

    def return_current_file_constants(self, file_name):
        clear_file_name = self.find_clear_file_name(file_name)
        constants = configurations.group_constants[clear_file_name]
        return constants

    def check_file_name(self, file_name):
        if '1_zovs' in file_name or '1_ZOVS' in file_name:
            return 'Имя файла ОК\n'
        elif '2_zovs' in file_name or '2_ZOVS' in file_name:
            return 'Имя файла ОК\n'
        elif '3_zovs' in file_name or '3_ZOVS' in file_name:
            return 'Имя файла ОК\n'
        elif '4_zovs' in file_name or '4_ZOVS' in file_name:
            return 'Имя файла ОК\n'
        elif '1_lovs' in file_name or '1_LOVS' in file_name:
            return 'Имя файла ОК\n'
        elif '2_lovs' in file_name or '2_LOVS' in file_name:
            return 'Имя файла ОК\n'
        elif '3_lovs' in file_name or '3_LOVS' in file_name:
            return 'Имя файла ОК\n'
        elif '4_lovs' in file_name or '4_LOVS' in file_name:
            return 'Имя файла ОК\n'
        else:
            return f'Ошибка в имени файла {file_name}\n'

    def find_clear_file_name(self, file_name):
        if '1_zovs' in file_name or '1_ZOVS' in file_name:
            return 'zovs_1'
        elif '2_zovs' in file_name or '2_ZOVS' in file_name:
            return 'zovs_2'
        elif '3_zovs' in file_name or '3_ZOVS' in file_name:
            return 'zovs_3'
        elif '4_zovs' in file_name or '4_ZOVS' in file_name:
            return 'zovs_4'
        elif '1_lovs' in file_name or '1_LOVS' in file_name:
            return 'lovs_1'
        elif '2_lovs' in file_name or '2_LOVS' in file_name:
            return 'lovs_2'
        elif '3_lovs' in file_name or '3_LOVS' in file_name:
            return 'lovs_3'
        elif '4_lovs' in file_name or '4_LOVS' in file_name:
            return 'lovs_4'
        else:
            return None

    def check_worksheet_names(self, worksheet_names):
        message = ''
        for worksheet_name in worksheet_names:
            if worksheet_name.lower() in configurations.words_to_skip:
                continue
            result = re.fullmatch(r'\d{2}[.]\d{2}[.]\s[-]\s\d{2}[.]\d{2}[.]', worksheet_name)
            if result == None:
                message += f'Ошибка в имени листа {worksheet_name}\n'
        if message == '':
            message = 'Имена листов ОК\n'
        return message

    #def check_group_numbers(self, work_book):
    #    message = ''
    #    for worksheet_name in work_book.sheetnames:
    #        worksheet = work_book[worksheet_name]
    #        for column in range(self.const_group_column, 25):
    #            group_number = worksheet.cell(row = self.const_group_row, column = column).value
    #            print(group_number)
    #            result = re.fullmatch(r'\bГруппа\s\d{3}', group_number)
    #            if result == None:
    #                message += f'Ошибка в номере группы {group_number} в листе {worksheet_name}'
    #    if message == '':
    #        message = 'Имена групп ОК'
    #    return message