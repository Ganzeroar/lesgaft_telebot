import time
from openpyxl import Workbook, load_workbook, utils
import glob

import db_funcs_for_subjects_db

# возвращает номер ряда, в котором в ячейках группы
def return_row_of_groups_number(work_sheet):
    for row in range(1, 10):
        if type(work_sheet.cell(row = row, column = 4).value) == str:
            if 'Группа' in work_sheet.cell(row = row, column = 4).value: 
                return row

# возвращает список номеров рядов в таблице, каждый из которых 
# содержит первую пару каждого дня. [номер ряда, в котором написаны все
# пары в понедельник в 9:45, номер ряда пар во вторник в 9:45 и т.д.]
# 
def return_list_of_row_of_time_of_first_lessons_of_all_days(work_sheet):
    list_of_times = []
    for row in range(1, 60):
        if type(work_sheet.cell(row = row, column = 3).value) == str:
            if '9:45' in work_sheet.cell(row = row, column = 3).value: 
                list_of_times.append(row)
            elif '09:45' in work_sheet.cell(row = row, column = 3).value: 
                list_of_times.append(row)
            elif '9.45' in work_sheet.cell(row = row, column = 3).value: 
                list_of_times.append(row)
            elif '09.45' in work_sheet.cell(row = row, column = 3).value: 
                list_of_times.append(row)
    return list_of_times

# возвращает список со списками дат, указанными слева от дней недели
# list_of_row_times_first_lessons = return_list_of_row_of_time_of_first_lessons_of_all_days?
def return_list_of_lists_of_dates(work_sheet):
    list_of_times_first_lessons =    []
    for row in range(1, 60):
        if type(work_sheet.cell(row = row, column = 3).value) == str:
            if '9:45' in work_sheet.cell(row = row, column = 3).value: 
                list_of_times_first_lessons.append(row)
            elif '09:45' in work_sheet.cell(row = row, column = 3).value: 
                list_of_times_first_lessons.append(row)
            elif '9.45' in work_sheet.cell(row = row, column = 3).value: 
                list_of_times_first_lessons.append(row)
            elif '09.45' in work_sheet.cell(row = row, column = 3).value: 
                list_of_times_first_lessons.append(row)
    returned_list = []
    for row in list_of_times_first_lessons:

        list_of_dates = [element.replace(' ', '') for element in work_sheet.cell(row = row, column = 1).value.rstrip().split('\n')]
        if list_of_dates[0] == '':
            del list_of_dates[0]

        final_list_of_dates = []
        print(list_of_dates)
        for date in list_of_dates:
            if date[-1] != '.':
                date = str(date) + '.'
                final_list_of_dates.append(date)
            else:
                final_list_of_dates.append(date)

        returned_list.append(final_list_of_dates)
    return returned_list

# проверяет соединены ли соседние ячейки
def isMerged(work_sheet, row, column):
    cell = work_sheet.cell(row, column)
    for mergedCell in work_sheet.merged_cells.ranges:
        if (cell.coordinate in mergedCell):
            return True
    return False

# берёт значение у соединённой чейки
def get_value_of_merged_call(work_sheet, row, column):
    cell = work_sheet.cell(row, column)
    for mergedCell in work_sheet.merged_cells.ranges:
        if (cell.coordinate in mergedCell):
            column_coor = list(mergedCell)[0][1]
            row_coor = list(mergedCell)[1][1]
            result_value = work_sheet.cell(row = row_coor, column = column_coor).value
            return result_value
    
def return_number_of_lessons_at_day(work_sheet, name_of_day):
    # Должна считать количество пар в день, но в 18 есть не у всех, потому
    # сейчас дефолтом вернёт 5
    first_lesson = '9:45'
    second_lesson = '11:30'
    third_lesson = '13:30'
    fourth_lesson = '15:15'
    fifth_lesson = '17:00'

    counter_of_number_of_lessons = 0

    for row in range(5, 50):
        if type(work_sheet.cell(row = row, column = 2).value) == str:
            if name_of_day in ''.join(work_sheet.cell(row = row, column = 2).value.split('\n')).lower():
                if work_sheet.cell(row = row, column = 3).value == first_lesson:
                    counter_of_number_of_lessons += 1
                if work_sheet.cell(row = row + 1, column = 3).value == second_lesson:
                    counter_of_number_of_lessons += 1
                if work_sheet.cell(row = row + 2, column = 3).value == third_lesson:
                    counter_of_number_of_lessons += 1
                if work_sheet.cell(row = row + 3, column = 3).value == fourth_lesson:
                    counter_of_number_of_lessons += 1
                if work_sheet.cell(row = row + 4, column = 3).value == fifth_lesson:
                    counter_of_number_of_lessons += 1
                #return counter_of_number_of_lessons
    #return False
    return 5

def return_full_data_of_day(work_sheet, name_for_db, list_of_day_dates, row_of_first_lesson_of_day, number_of_lessons_at_day, row_of_groups_number):
    for date in list_of_day_dates:
        for num_of_lesson in range(number_of_lessons_at_day):
            time = work_sheet.cell(row = row_of_first_lesson_of_day + num_of_lesson, column = 3).value
            if type(time) != str:
                continue
            if time[0] == '0':
                time = time.replace('0', '')
            time = time.replace('.', ':')
            
            for group_column in range(4, 25):
                group_cell_value = work_sheet.cell(row = row_of_groups_number, column = group_column).value
                if type(group_cell_value) == str and 'Группа' in group_cell_value:
                    try:
                        if '  ' in group_cell_value:
                            group_cell_value = group_cell_value.replace('  ', ' ')
                        if ' ' in group_cell_value:
                            group_cell_value = group_cell_value.replace(' ', '_')
                        subject = work_sheet.cell(row = row_of_first_lesson_of_day + num_of_lesson, column = group_column).value
                        if isMerged(work_sheet, row_of_first_lesson_of_day + num_of_lesson, group_column):
                            merged_subject = get_value_of_merged_call(work_sheet, row_of_first_lesson_of_day + num_of_lesson, group_column)
                            db_funcs_for_subjects_db.save_subj(name_for_db, date, time, group_cell_value, merged_subject.replace("\n"," "))
                        elif type(subject) == str:
                            db_funcs_for_subjects_db.save_subj(name_for_db, date, time, group_cell_value, subject.replace("\n"," "))
                        elif subject == None:
                            db_funcs_for_subjects_db.save_subj(name_for_db, date, time, group_cell_value, 'Нет предмета')
                    except:
                        print(group_cell_value)
                        print(name_for_db)
                        print('ERROR Появилась новая группа')
                        return False
                        
def return_list_of_groups(work_sheet, row_of_groups_number):
    list_of_groups = []
    for column in range(4, 25):
        group_cell_value = work_sheet.cell(row = row_of_groups_number, column = column).value
        if type(group_cell_value) == str and 'Группа' in group_cell_value:
            if '  ' in group_cell_value:
                group_cell_value = group_cell_value.replace('  ', ' ')
            if ' ' in group_cell_value:
                group_cell_value = group_cell_value.replace(' ', '_')
            list_of_groups.append(group_cell_value)
    return list_of_groups

def return_list_of_times(work_sheet, number_of_lessons_at_day):
    first_lesson = '9:45'
    second_lesson = '11:30'
    third_lesson = '13:30'
    fourth_lesson = '15:15'
    fifth_lesson = '17:00'

    if number_of_lessons_at_day == 1:
        return [first_lesson]
    elif number_of_lessons_at_day == 2:
        return [first_lesson, second_lesson]
    elif number_of_lessons_at_day == 3:
        return [first_lesson, second_lesson, third_lesson]
    elif number_of_lessons_at_day == 4:
        return [first_lesson, second_lesson,third_lesson, fourth_lesson]
    elif number_of_lessons_at_day == 5:
        return [first_lesson, second_lesson, third_lesson, fourth_lesson, fifth_lesson]
    
def return_db_name(file_name):
    if 'zovs_1_kurs' in file_name:
        return 'zovs_1_kurs'
    elif 'zovs_2_kurs' in file_name:
        return 'zovs_2_kurs'
    elif 'zovs_3_kurs' in file_name:
        return 'zovs_3_kurs'
    elif 'zovs_4_kurs' in file_name:
        return 'zovs_4_kurs'
    elif 'lovs_1_kurs' in file_name:
        return 'lovs_1_kurs'
    elif 'lovs_2_kurs' in file_name:
        return 'lovs_2_kurs'
    elif 'lovs_3_kurs' in file_name:
        return 'lovs_3_kurs'
    elif 'lovs_4_kurs' in file_name:
        return 'lovs_4_kurs'

def create_groups_in_db(work_book, work_sheet_for_create_groups, db_name):
    main_work_sheet = work_book[work_sheet_for_create_groups]
    row_of_groups_number = return_row_of_groups_number(main_work_sheet)
    list_of_groups = return_list_of_groups(main_work_sheet, row_of_groups_number)
    db_funcs_for_subjects_db.save_groups(db_name, list_of_groups)
    
def pars_files_create_dbfiles():
    work_files = glob.glob('time_tables/full_time_undergraduate/*.xlsx')
    for work_file in work_files:
        print(work_file)
        db_name = return_db_name(work_file)
        db_funcs_for_subjects_db.create_db(db_name)
        work_book = load_workbook(work_file)
        for work_sheet_for_create_groups in work_book.sheetnames:
            create_groups_in_db(work_book, work_sheet_for_create_groups, db_name)
            break
        for work_sheet in work_book.sheetnames:
            main_work_sheet = work_book[work_sheet]
            list_of_dates = return_list_of_lists_of_dates(main_work_sheet)
            list_of_times = return_list_of_times(main_work_sheet, 5)
            db_funcs_for_subjects_db.save_dates_and_times(db_name, list_of_dates, list_of_times)

            days_names = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота']
            for i in range(6):
                day_name = days_names[i]
                list_of_day_dates = return_list_of_lists_of_dates(main_work_sheet)[i]
                row_of_first_lesson_of_day = return_list_of_row_of_time_of_first_lessons_of_all_days(main_work_sheet)[i]
                number_of_lessons_at_day = return_number_of_lessons_at_day(main_work_sheet, day_name)
                row_of_groups_number = return_row_of_groups_number(main_work_sheet)

                return_full_data_of_day(main_work_sheet, db_name, list_of_day_dates, row_of_first_lesson_of_day, number_of_lessons_at_day, row_of_groups_number)
            print('Done WS ' + str(work_sheet))
        print('Done ' + str(work_file))
    print('All is Done')
    return True
if __name__ == "__main__":
    pars_files_create_dbfiles()
        
        

