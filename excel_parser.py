from openpyxl import Workbook, load_workbook, utils
import glob
import datetime

import db_funcs_for_subjects_db
import configurations
import excel_validator
import main


class Excel_parser():

    const_dates_column = 1
    const_time_column = 3
    const_quantity_of_rows = 50
    const_first_group_column = 4

    def run_parser(self, route):
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        print(work_files)
        validator = excel_validator.Excel_validator()
        result_message = validator.run_validator_for_excel_parser(route)
        if 'Ошибка' in result_message:
            return result_message

        for work_file in work_files:
            print('file = ' + work_file)
            self.parse_work_file(work_file)

        response_is_null_in_db = db_funcs_for_subjects_db.check_is_null_in_db()
        if response_is_null_in_db != 'OK':
            main.send_custom_message_to_user(
            206171081, f'null в БД {response_is_null_in_db}')
        else:
            print('DB HAS NOT NULL')

    def parse_work_file_using_name(self, name, route):
        print('Парсер запущен на ' + name)

        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        for work_file in work_files:
            if name in work_file:
                db_name = self.return_db_name(work_file.lower())
                work_book = load_workbook(work_file)
                db_funcs_for_subjects_db.drop_db(db_name)
                db_funcs_for_subjects_db.create_db(db_name)
                self.create_groups_in_db(work_book, db_name)

                for ws in work_book.sheetnames:
                    work_sheet = work_book[ws]
                    ws_name = str(work_sheet)
                    if self.is_reason_to_skip(ws_name):
                        print('skipped' + ws_name)
                        continue
                    print('ws start ' + str(work_sheet))
                    self.create_dates_and_times_in_db(work_sheet, db_name)
                    self.parse_work_sheet(work_sheet, db_name)

    def parse_work_file(self, work_file):
        db_name = self.return_db_name(work_file.lower())
        db_funcs_for_subjects_db.drop_db(db_name)
        db_funcs_for_subjects_db.create_db(db_name)

        work_book = load_workbook(work_file)
        self.create_groups_in_db(work_book, db_name)

        for ws in work_book.sheetnames:
            work_sheet = work_book[ws]
            ws_name = str(work_sheet)
            if self.is_reason_to_skip(ws_name):
                print('skipped' + ws_name)
                continue
            print('sheet = ' + str(work_sheet))
            self.create_dates_and_times_in_db(work_sheet, db_name)
            self.parse_work_sheet(work_sheet, db_name)

    def create_groups_in_db(self, work_book, db_name):
        for ws in work_book.sheetnames:
            work_sheet = work_book[ws]
            ws_name = str(work_sheet)
            if self.is_reason_to_skip(ws_name):
                print('skipped' + ws_name)
                continue
            print('create groups in ' + str(work_sheet))
            first_group_name = self.return_first_group_name(db_name)
            list_of_groups = self.return_all_groups_names(
                work_sheet, first_group_name, db_name)
            db_funcs_for_subjects_db.save_groups(db_name, list_of_groups)
            return

    def create_dates_and_times_in_db(self, work_sheet, db_name):
        dates_column = self.const_dates_column
        time_column = self.const_time_column
        times = []
        for row in range(1, self.const_quantity_of_rows):
            time_cell = str(work_sheet.cell(row=row, column=time_column).value)
            if time_cell != None and self.is_time(time_cell):
                time_value = self.format_time(time_cell)
                if time_value == '9:45':
                    times = []
                    times.append(time_value)
                    dates = self.get_dates(work_sheet, row, dates_column, db_name)
                elif time_value == '11:30' or time_value == '13:30' or time_value == '15:15':
                    times.append(time_value)
                elif time_value == '17:00':
                    times.append(time_value)
                    next_cell = str(work_sheet.cell(
                        row=row + 1, column=time_column).value)
                    next_cell = self.format_time(
                        next_cell) if self.is_time(next_cell) else False

                    after_next_cell = str(work_sheet.cell(
                        row=row + 2, column=time_column).value)
                    after_next_cell = self.format_time(
                        after_next_cell) if self.is_time(after_next_cell) else False

                    if self.is_time(next_cell) or self.is_time(after_next_cell):
                        if next_cell == '9:45' or after_next_cell == '9:45' and next_cell != '18:40':
                            self.save_dates_and_times(db_name, dates, times)
                            continue
                        elif next_cell == '18:40' or after_next_cell == '18:40':
                            continue
                    elif next_cell == False and after_next_cell == False:
                        self.save_dates_and_times(db_name, dates, times)
                elif time_value == '18:40':
                    times.append(time_value)
                    self.save_dates_and_times(db_name, dates, times)

    def parse_work_sheet(self, work_sheet, db_name):
        time_column = self.const_time_column
        dates_column = self.const_dates_column

        first_group_name = self.return_first_group_name(db_name)
        groups_columns = self.return_columns_numbers_of_all_groups_cells(
            work_sheet, first_group_name)
        first_row = self.find_row_of_first_lesson(work_sheet)
        if first_row == None:
            return False
        groups_row = self.find_number_of_groups_cell_row(
            work_sheet, first_group_name)
        times = ['9:45', '11:30', '13:30', '15:15', '17:00', '18:40']
        for column in groups_columns:
            for row in range(first_row, self.const_quantity_of_rows):
                subject = work_sheet.cell(row=row, column=column).value
                if self.is_merged(work_sheet, row, column):
                    subject = self.get_value_of_merged_call(
                        work_sheet, row, column)
                if subject == None:
                    subject = 'нет предмета'
                time_cell = work_sheet.cell(row=row, column=time_column).value
                if time_cell == None:
                    continue
                time = self.format_time(str(time_cell))
                if time in times:
                    dates = work_sheet.cell(row=row, column=dates_column).value
                    if self.is_merged(work_sheet, row, dates_column):
                        dates = self.get_value_of_merged_call(
                            work_sheet, row, dates_column)
                    if self.is_merged(work_sheet, row, dates_column) == False and self.is_merged(work_sheet, row-1, dates_column):
                        dates = self.get_value_of_merged_call(
                            work_sheet, row-1, dates_column)
                    group_name = work_sheet.cell(
                        row=groups_row, column=column).value
                    if bool(group_name) == False:
                        continue
                    group_name = self.format_group_name(group_name)
                    undegrdaduate_timetables = ['zovs_1', 'zovs_2', 'zovs_3',
                                                'zovs_4', 'lovs_1', 'lovs_2', 'lovs_3', 'lovs_4']

                    if db_name in undegrdaduate_timetables:

                        if self.is_merged(work_sheet, groups_row-1, column):
                            group_number = self.format_group_name(
                                self.get_value_of_merged_call(work_sheet, groups_row-1, column))
                        else:
                            group_number = self.format_group_name(
                                work_sheet.cell(row=groups_row-1, column=column).value)
                        group_name = group_name + '_' + group_number

                    if not db_funcs_for_subjects_db.is_group_exist(group_name, db_name):
                        db_funcs_for_subjects_db.save_group(
                            db_name, group_name)
                    self.save_subj_in_db(
                        db_name, dates, time, group_name, subject)
                else:
                    print(
                        f'Какая-то ошибка в эксель парсере в колонке {column}, ряде {row}, расписании {db_name}, холсте {work_sheet}, предмете {subject}, времени {time}')
        print('ws finished')

    def find_number_of_groups_cell_row(self, work_sheet, first_group_name):
        first_group_column = self.const_first_group_column
        for row in range(1, 10):
            viewed_cell = str(work_sheet.cell(
                row=row, column=first_group_column).value)
            if type(viewed_cell) == str:
                viewed_cell = self.format_group_name(viewed_cell)
                if first_group_name in viewed_cell:
                    return row

    def validate_date(self, date, db_name):
        dates_in_db = db_funcs_for_subjects_db.get_dates(db_name)
        for elem in dates_in_db:
            if elem [0] == date:
                current_date = datetime.datetime.strptime(f"2021-{date[3:5]}-{date[0:2]}","%Y-%m-%d")
                future_date = datetime.timedelta(days=7)
                days_to_add = current_date + future_date
                date_next_week = days_to_add.strftime("%d.%m.%Y.")
                date_next_week_formatted = date_next_week[0:2] + '.' + date_next_week[3:5] + '.'
                return date_next_week_formatted

        return date

    def return_all_groups_names(self, work_sheet, first_group_name, db_name):
        undegrdaduate_timetables = ['zovs_1', 'zovs_2', 'zovs_3',
                                    'zovs_4', 'lovs_1', 'lovs_2', 'lovs_3', 'lovs_4']
        groups_names = []
        row_number = self.find_number_of_groups_cell_row(
            work_sheet, first_group_name)
        first_group_column = self.const_first_group_column
        for column in range(first_group_column, 25):
            group_cell = work_sheet.cell(row=row_number, column=column).value
            if type(group_cell) == str:
                group_cell = self.format_group_name(group_cell)
                if db_name in undegrdaduate_timetables:
                    if self.is_merged(work_sheet, row_number-1, column):
                        group_number = self.get_value_of_merged_call(
                            work_sheet, row_number-1, column)
                    else:
                        group_number = work_sheet.cell(
                            row=row_number-1, column=column).value

                    group_number_formatted = self.format_group_name(
                        group_number)
                    group_cell = group_cell + '_' + group_number_formatted

                groups_names.append(group_cell)
        return groups_names

    def return_columns_numbers_of_all_groups_cells(self, work_sheet, first_group_name):
        columns_numbers_of_all_groups_cells = []
        row_number = self.find_number_of_groups_cell_row(   
            work_sheet, first_group_name)
        first_group_column = self.const_first_group_column
        for column in range(first_group_column, 25):
            group_cell = work_sheet.cell(row=row_number, column=column).value
            if type(group_cell) == str:
                group_cell = self.format_group_name(group_cell)
                columns_numbers_of_all_groups_cells.append(column)
        return columns_numbers_of_all_groups_cells

    def is_merged(self, work_sheet, row, column):
        cell = work_sheet.cell(row, column)
        for mergedCell in work_sheet.merged_cells.ranges:
            if (cell.coordinate in mergedCell):
                return True
        return False

    def is_time(self, time):
        #TODO убрать так как это делает валидатор
        times = ['9:45', '09:45', '9.45', '09.45', '11:30', '11.30', '13:30',
                 '13.30', '15:15', '15.15', '17:00', '17.00', '18:40', '18.40', '9:45:00',
                 '09:45:00', '9.45:00', '09.45:00', '11:30:00', '11.30:00', '13:30:00',
                 '13.30:00', '15:15:00', '15.15:00', '17:00:00', '17.00:00', '18:40:00', '18.40:00']
        if time in times:
            return True
        else:
            return False

    def format_time(self, time):
        time = str(time)
        #TODO убрать так как это делает валидатор
        # в расписании чаще всего косячат тут
        first_times = ['9:45', '09:45', '9.45', '09.45',
                       '9:45:00', '09:45:00', '9.45:00', '09.45:00']
        if time in first_times:
            time = '9:45'
        if '.' in time:
            time = time.replace('.', ':')
        if ';' in time:
            time = time.replace(';', ':')
        if len(time) >= 8 and time[-1] == '0' and time[-2] == '0' and time[-3] == ':':
            if len(time) == 8:
                time = time[:5]
            elif len(time) == 7:
                time = time[:4]
        return time

    def is_reason_to_skip(self, worksheet_name):
        month_to_skip = configurations.month_to_skip

        for num in range(0, -15, -1):
            if worksheet_name[num-2:num].isdigit():
                if worksheet_name[num-2:num] in month_to_skip:
                    return True
                else:
                    break
        words_to_skip = configurations.words_to_skip
        lower_ws_name = worksheet_name.lower()
        for word in words_to_skip:
            if word in lower_ws_name:
                return True
        return False

    def save_dates_and_times(self, db_name, dates, times):
        for date in dates:
            date = self.validate_date(date, db_name)
            for time in times:
                db_funcs_for_subjects_db.save_date_and_time(
                    db_name, date, time)

    def format_dates(self, dates):
        #TODO убрать так как это делает валидатор

        dates = dates.replace(' ', '\n')
        list_of_dates = [element.replace(' ', '')
                         for element in dates.rstrip().split('\n')]
        formatted_list_of_dates = []
        for date in list_of_dates:
            if date == '':
                continue
            elif date[-1] != '.':
                date = date + '.'
            if len(date) >= 12:
                first_date = date[:6]
                second_date = date[6:]
                formatted_list_of_dates.append(first_date)
                formatted_list_of_dates.append(second_date)
                continue
            if '..' in date:
                date = date.replace('..', '.')
            formatted_list_of_dates.append(date)
        return formatted_list_of_dates

    def get_dates(self, work_sheet, row, dates_column, db_name):
        dates = work_sheet.cell(row=row, column=dates_column).value
        if self.is_merged(work_sheet, row, dates_column):
            dates = self.get_value_of_merged_call(
                work_sheet, row, dates_column)
        dates = self.format_dates(dates)
        dates = self.validate_date(dates, db_name)
        return dates

    def get_value_of_merged_call(self, work_sheet, row, column):
        cell = work_sheet.cell(row, column)
        cell_value = []
        if cell.value != None:
            return cell.value
        for mergedCell in work_sheet.merged_cells.ranges:
            if cell.coordinate in mergedCell:
                cell_value.append(mergedCell)
                #column_coor = list(mergedCell)[0][1]
                #row_coor = list(mergedCell)[1][1]
                #result_value = work_sheet.cell(
                #    row=row_coor, column=column_coor).value
        result_value = work_sheet.cell(cell_value[0].min_row, cell_value[0].min_col).value
        return result_value

    def find_row_of_first_lesson(self, work_sheet):
        time_column = self.const_time_column
        #TODO убрать так как это делает валидатор

        times = ['9:45', '09:45', '9.45', '09.45',
                 '9:45:00', '09:45:00', '9.45:00', '09.45:00']
        for row in range(1, 10):
            time_cell = str(work_sheet.cell(row=row, column=time_column).value)
            if time_cell in times:
                return row

    def save_subj_in_db(self, db_name, dates, time, group_name, subject):
        dates = self.format_dates(dates)

        for date in dates:
            db_funcs_for_subjects_db.save_subj(
                db_name, date, time, group_name, subject)

    group_names = ['гр_1', 'гр_2', 'гр_3', 'гр_4', 'гр_5', 'гр_6', 'гр_7',
                   'гр_8', 'гр_9', 'гр_10', 'гр_11', 'гр_12', 'гр_13',
                   'гимнастика_плавание_футбол',
                   'лёгкая_атлетика_спортивные_игры',
                   'велоспорт_водномоторный_и_парусный_спорт_гребной_спорт_атлетизм_бокс_борьба_фехтование_нвс',
                   'хоккей_керлинг_биатлон_лыжный_спорт_конькобежный_спорт_и_фк',
                   'профессиональное_образование_в_сфере_физической_культуры_и_спорта',
                   'комплексное_научное_обеспечение_спортивной_подготовки',
                   'физкультурно_оздоровительная_работа',
                   'комплексная_реабилитация_в_физической_культуре_и_спорте',
                   'медико_биологическое_сопровождение_физической_культуры_и_спорта',
                   'менеджмент_направленность_профиль_менеджмент_в_спорте',
                   'туризм_направленность_профиль_туристская_деятельность_с_сфере_физической_культуры_и_спорта',
                   'журналистика_направленность_профиль_спортивная_журналистика',
                   'государственное_и_муниципальное_управление_направленность_профиль_государственное_и_муниципальное_управление_в_отрасли_физической_культуры_и_спорта',
                   'менеджмент_направленность_профиль_менеджмент_в_спорте',
                   'туризм_направленность_профиль_туристская_деятельность_с_сфере_физической_культуры_и_спорта',
                   'журналистика_направленность_профиль_спортивная_журналистика',
                   'адаптивное_физическое_воспитание_в_системе_образования_обучающихся_с_овз',
                   'спортивная_подготовка_лиц_с_овз_включая_инвалидов',
                   'физическая_реабилитация',
                   'педагогическая_гидрореабилитация',
                   'современная_хореография_и_танцы_в_коррекции_нарушений_у_лиц_с_овз',
                   'адаптивное_физическое_воспитание_в_системе_образования_обучающихся_с_ограниченными_возможностями_здоровья',
                   'технологии_профилактики_и_коррекции_аддиктивного_поведения',
                   'спортивная_подготовка_лиц_с_ограниченными_возможностями_здоровья_включая_инвалидов',
                   'педагогическая_гидрореабилитация',
                   'физическая_реабилитация',
                   'современная_хореография_и_танцы_в_коррекции_нарушений_у_лиц_с_отклонениями_в_состоянии_здоровья',
                   'менеджмент',
                   'международные_отношения',
                   'журналистика',
                   'туризм',
                   'сервис',
                   'реклама и связи с общественностью'
                   ]

    def return_db_name(self, file_name):
        if '1_zovs' in file_name:
            return 'zovs_1'
        elif '2_zovs' in file_name:
            return 'zovs_2'
        elif '3_zovs' in file_name:
            return 'zovs_3'
        elif '4_zovs' in file_name:
            return 'zovs_4'
        elif '1_lovs' in file_name:
            return 'lovs_1'
        elif '2_lovs' in file_name:
            return 'lovs_2'
        elif '3_lovs' in file_name:
            return 'lovs_3'
        elif '4_lovs' in file_name:
            return 'lovs_4'
        elif 'imst_1_kurs' in file_name:
            return 'imst_1_kurs'
        elif 'imst_2_kurs' in file_name:
            return 'imst_2_kurs'
        elif 'imst_3_kurs' in file_name:
            return 'imst_3_kurs'
        elif 'imst_4_kurs' in file_name:
            return 'imst_4_kurs'
        elif 'magistracy_fk_full_time_1_kurs' in file_name:
            return 'magistracy_fk_full_time_1_kurs'
        elif 'magistracy_fk_full_time_2_kurs' in file_name:
            return 'magistracy_fk_full_time_2_kurs'
        elif 'magistracy_imst_full_time_1_kurs' in file_name:
            return 'magistracy_imst_full_time_1_kurs'
        elif 'magistracy_imst_full_time_2_kurs' in file_name:
            return 'magistracy_imst_full_time_2_kurs'
        elif 'magistracy_afk_full_time_1_kurs' in file_name:
            return 'magistracy_afk_full_time_1_kurs'
        elif 'magistracy_afk_full_time_2_kurs' in file_name:
            return 'magistracy_afk_full_time_2_kurs'

    def format_group_name(self, group_name):
        if bool(group_name) == False:
            return
        group_name = group_name.lower().rstrip()
        unnecessary_symbols = [';', ':', '(', ')', '"', '.', ',']
        symbols_for_change = ['\n', ' ', '-']
        numbers_for_delete_before = [
            '380404', '380402', '430402', '420402', '410405']
        numbers_for_delete_after = [
            '380302', '430302', '430301', '410305', '420301', '420302']
        if 'гр.' in group_name:
            first = group_name[:-3]
            second = group_name[2:]
            group_name = second + '_' + first
        while '  ' in group_name:
            group_name = group_name.replace('  ', ' ')
        for symbol in unnecessary_symbols:
            if symbol in group_name:
                group_name = group_name.replace(symbol, '')
        for symbol in symbols_for_change:
            if symbol in group_name:
                group_name = group_name.replace(symbol, '_')
        for num in numbers_for_delete_before:
            if num in group_name:
                group_name = group_name[6:]
        for num in numbers_for_delete_after:
            if num in group_name:
                group_name = group_name[:-6]
        while group_name[0] == '_':
            group_name = group_name[1:]
        if group_name[-1] == '_':
            group_name = group_name[:-1]
        while '/' in group_name:
            group_name = group_name.replace('/', '_')
        while '__' in group_name:
            group_name = group_name.replace('__', '_')
        if group_name[-1] == '_':
            group_name = group_name[:-1]

        return group_name
        
    def return_first_group_name(self, db_name):
        if db_name == 'magistracy_fk_full_time_1_kurs':
            first_group_name = 'гимнастика'
        elif db_name == 'magistracy_fk_full_time_2_kurs':
            first_group_name = 'гимнастика'
        elif db_name == 'magistracy_imst_full_time_1_kurs':
            first_group_name = 'туризм_направленность_профиль_туристская_деятельность_с_сфере_физической_культуры_и_спорта'
        elif db_name == 'magistracy_imst_full_time_2_kurs':
            first_group_name = 'менеджмент_8_направленность_профиль_менеджмент_в_спорте'
        elif db_name == 'magistracy_afk_full_time_1_kurs':
            first_group_name = 'адаптивное_физическое_воспитание_в_системе_образования_обучающихся_с_овз'
        elif db_name == 'magistracy_afk_full_time_2_kurs':
            first_group_name = 'адаптивное_физическое_воспитание_в_системе_образования_обучающихся_с_овз'
        elif db_name == 'zovs_1':
            first_group_name = 'атлетизм_тхэквондо_тхэквондо_антидопинг'
            #first_group_name = 'группа_113'
        elif db_name == 'zovs_2':
            first_group_name = 'пауэрлифтинг_гиревой_спорт_бодибилдинг_тяжелая_атлетика_фехтование_фехтование_антидопинг'
            #first_group_name = 'группа_212'
        elif db_name == 'zovs_3':
            first_group_name = 'атлетизм_пауэрлифтинг_гиревой_спорт_бодибилдинг_тяжелая_атлетика_бокс'
            #first_group_name = 'группа_312'
        elif db_name == 'zovs_4':
            first_group_name = 'самбо_атлетизм'
            #first_group_name = 'группа_411'
        elif db_name == 'lovs_1':
            first_group_name = 'худ_гимн_худ_гимн_антидопинг'
            #first_group_name = 'группа_101'
        elif db_name == 'lovs_2':
            first_group_name = 'худ_гимн_худ_гимн_антидопинг'
            #first_group_name = 'группа_201'
        elif db_name == 'lovs_3':
            first_group_name = 'худ_гимн'
            #first_group_name = 'группа_301'
        elif db_name == 'lovs_4':
            first_group_name = 'худ_гимн'
            #first_group_name = 'группа_401'
        elif db_name == 'imst_1_kurs':
            first_group_name = 'менеджмент'
        elif db_name == 'imst_2_kurs':
            first_group_name = 'менеджмент'
        elif db_name == 'imst_3_kurs':
            first_group_name = 'менеджмент'
        elif db_name == 'imst_4_kurs':
            first_group_name = 'менеджмент'
        return first_group_name


class Excel_parser_undergraduate_imst(Excel_parser):
    const_first_group_column = 7
    const_time_column = 6
    const_dates_column = 4


def run_undergraduate_parser():
    parser = Excel_parser()
    parser.run_parser('full_time_undergraduate')


def run_undergraduate_imst_parser():
    parser = Excel_parser_undergraduate_imst()
    parser.run_parser('full_time_undergraduate/imst')


def run_magistracy_fk_parser():
    parser = Excel_parser()
    parser.run_parser('full_time_magistracy_fk')


def run_magistracy_afk_parser():
    parser = Excel_parser()
    parser.run_parser('full_time_magistracy_afk')


def run_magistracy_imst_parser():
    parser = Excel_parser()
    parser.run_parser('full_time_magistracy_imst')

    parser = Excel_parser()
    parser.run_parser('full_time_magistracy_fk')
    parser = Excel_parser()
    parser.run_parser('full_time_magistracy_imst')
    parser = Excel_parser()
    parser.run_parser('full_time_magistracy_afk')
    parser = Excel_parser()
    parser.run_parser('full_time_undergraduate')
    parser = Excel_parser_undergraduate_imst()
    parser.run_parser('full_time_undergraduate/imst')


def run_all_parsers():
    parser = Excel_parser()
    parser.run_parser('full_time_magistracy_fk')
    parser = Excel_parser()
    parser.run_parser('full_time_magistracy_imst')
    parser = Excel_parser()
    parser.run_parser('full_time_magistracy_afk')
    parser = Excel_parser()
    parser.run_parser('full_time_undergraduate')
    parser = Excel_parser_undergraduate_imst()
    parser.run_parser('full_time_undergraduate/imst')


if __name__ == "__main__":
    run_undergraduate_parser()

#оставить все строчки на 17:00
#добавить строчку на 18:40 только когда есть пары
#парсер парсит - видит пустую строку - оповещает об этом составителя расписания
