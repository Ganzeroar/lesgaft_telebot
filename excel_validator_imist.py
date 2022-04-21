from openpyxl import load_workbook
import glob
import re
import configurations
import os

from file_not_valid_exception import File_not_valid
from excel_validator_main import Excel_validator

class Excel_validator_imist(Excel_validator):

    def run_validator(self, route):
        work_files = glob.glob(f'time_tables/{route}/*.xlsx')
        for work_file_name in work_files:
            try:
                self.check_file_name(work_file_name)
                work_book = load_workbook(work_file_name)
                self.check_worksheet_names(work_book.sheetnames)
                self.check_structure(work_book, work_file_name)
            finally:
                path = os.path.join(os.path.abspath(os.path.dirname(__file__)), work_file_name)
                os.remove(path)
        return f'{work_file_name} валиден'

    def check_structure(self, work_book, work_file_name):
        for worksheet_name in work_book.sheetnames:
            if self.is_reason_to_skip(worksheet_name) == True:
                continue
            worksheet = work_book[worksheet_name]
            self.check_group_struct(worksheet, work_file_name, worksheet_name)
            self.check_date_struct(worksheet, work_file_name, worksheet_name)
            self.check_day_struct(worksheet, work_file_name, worksheet_name)
            self.check_time_struct(worksheet, work_file_name, worksheet_name)
        return 'Структура ОК\n'
    
    def check_group_struct(self, worksheet, work_file_name, worksheet_name):
        constants = self.return_current_file_constants(work_file_name)
        const_group_row = constants['group_row']
        const_first_group_first_column = constants['first_group_first_column']
        const_first_group_last_column = constants['first_group_last_column']
        const_first_group_number = constants['first_group_number']
        const_second_group_first_column = constants['second_group_first_column']
        const_second_group_last_column = constants['second_group_last_column']
        const_second_group_number = constants['second_group_number']
        const_third_group_first_column = constants['third_group_first_column']
        const_third_group_last_column = constants['third_group_last_column']
        const_third_group_number = constants['third_group_number']
        
        for column in range(const_first_group_first_column, const_first_group_last_column + 1):
            viewed_group_cell = worksheet.cell(row = const_group_row, column = column)
            if self.is_merged(worksheet, viewed_group_cell):
                viewed_group_value = self.get_merged_cell_value(worksheet, viewed_group_cell)
                if viewed_group_value != const_first_group_number:
                    raise File_not_valid(f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')
            else:
                raise File_not_valid(f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')
        
        for column in range(const_second_group_first_column, const_second_group_last_column + 1):
            viewed_group_cell = worksheet.cell(row = const_group_row, column = column)
            if self.is_merged(worksheet, viewed_group_cell):
                viewed_group_value = self.get_merged_cell_value(worksheet, viewed_group_cell)
                if viewed_group_value != const_second_group_number:
                    raise File_not_valid(f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')
            else:
                raise File_not_valid(f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')
        
        for column in range(const_third_group_first_column, const_third_group_last_column + 1):
            viewed_group_cell = worksheet.cell(row = const_group_row, column = column)
            if self.is_merged(worksheet, viewed_group_cell):
                viewed_group_value = self.get_merged_cell_value(worksheet, viewed_group_cell)
                if viewed_group_value != const_third_group_number:
                    raise File_not_valid(f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')
            else:
                raise File_not_valid(f'Ошибка в структуре группы в {viewed_group_cell.coordinate} в листе {worksheet_name}')        

    def find_clear_file_name(self, file_name):
        if '1_imist' in file_name:
            return 'imist_1'
        elif '2_imist' in file_name:
            return 'imist_2'
        elif '3_imist' in file_name:
            return 'imist_3'
        elif '4_imist' in file_name:
            return 'imist_4'
        else:
            return None

    def return_current_file_constants(self, work_file_name):
        clear_file_name = self.find_clear_file_name(work_file_name)
        constants = configurations.timetable_constants[clear_file_name]
        return constants
    
    #TODO потенциально можно вынести в константы и объединить с другими
    def check_file_name(self, work_file_name):
        if '1_imist' in work_file_name:
            return 'Имя файла ОК\n'
        elif '2_imist' in work_file_name:
            return 'Имя файла ОК\n'
        elif '3_imist' in work_file_name:
            return 'Имя файла ОК\n'
        elif '4_imist' in work_file_name:
            return 'Имя файла ОК\n'
        else:
            raise File_not_valid(f'Ошибка в имени файла {work_file_name}')