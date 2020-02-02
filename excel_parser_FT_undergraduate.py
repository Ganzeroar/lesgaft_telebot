import time
from openpyxl import Workbook, load_workbook, utils
import glob

import db_funcs_for_subjects_db

# константное значение
# пока что во всех расписаниях колонка первых групп одинакова 
const_first_group_column = 4
const_time_column = 3
const_dates_column = 1
const_quantity_of_rows = 50

def find_number_of_groups_cell_row(work_sheet):
    first_group_column = const_first_group_column
    for row in range(1, 10):
        viewed_cell = work_sheet.cell(row = row, column = first_group_column).value 
        if type(viewed_cell) == str and 'Группа' in viewed_cell:
            return row

def return_columns_numbers_of_all_groups_cells(work_sheet):
    columns_numbers_of_all_groups_cells = []
    row_number = find_number_of_groups_cell_row(work_sheet)
    first_group_column = const_first_group_column
    for column in range(first_group_column, 25):
        group_cell = work_sheet.cell(row = row_number, column = column).value
        if type(group_cell) == str and 'Группа' in group_cell:
            columns_numbers_of_all_groups_cells.append(column)
    return columns_numbers_of_all_groups_cells

def return_all_groups_names(work_sheet):
    groups_names = []
    row_number = find_number_of_groups_cell_row(work_sheet)
    first_group_column = const_first_group_column
    for column in range(first_group_column, 25):
        group_cell = work_sheet.cell(row = row_number, column = column).value
        if type(group_cell) == str and 'Группа' in group_cell:
            group_name = format_group_name(group_cell)
            groups_names.append(group_name)
    return groups_names


def find_row_of_first_lesson(work_sheet):
    time_column = const_time_column
    times = ['9:45', '09:45', '9.45', '09.45']
    for row in range(1, 10):
        time_cell = work_sheet.cell(row = row, column = time_column).value
        if time_cell in times:
            return row 

def is_merged(work_sheet, row, column):
    cell = work_sheet.cell(row, column)
    for mergedCell in work_sheet.merged_cells.ranges:
        if (cell.coordinate in mergedCell):
            return True
    return False

def get_value_of_merged_call(work_sheet, row, column):
    cell = work_sheet.cell(row, column)
    for mergedCell in work_sheet.merged_cells.ranges:
        if (cell.coordinate in mergedCell):
            column_coor = list(mergedCell)[0][1]
            row_coor = list(mergedCell)[1][1]
            result_value = work_sheet.cell(row = row_coor, column = column_coor).value
            return result_value

def format_group_name(group_name):
    if '  ' in group_name:
        group_name = group_name.replace('  ', ' ')
    if ' ' in group_name:
        group_name = group_name.replace(' ', '_')
    if '\n' in group_name:
        group_name = group_name.replace('\n', '')
    group_name = group_name.rstrip()
    return group_name

def format_dates(dates):
    list_of_dates = [element.replace(' ', '') for element in dates.rstrip().split('\n')]
    formatted_list_of_dates = []
    for date in list_of_dates:
        if date == '':
            pass
        elif date[-1] != '.':
            date = date + '.'
            formatted_list_of_dates.append(date)
        else:
            formatted_list_of_dates.append(date)
    return formatted_list_of_dates

def format_time(time):
    # в расписании чаще всего косячат тут
    first_times = ['9:45', '09:45', '9.45', '09.45']
    if time in first_times:
        time = '9:45'
    elif '.' in time:
        time = time.replace('.', ':')
    return time

def is_time(time):
    times = ['9:45', '09:45', '9.45', '09.45', '11:30', '11.30', '13:30', 
    '13.30', '15:15', '15.15', '17:00', '17.00', '18:40', '18.40']
    if time in times:
        return True
    else:
        return False

def save_subj_in_db(db_name, dates, time, group_name, subject):
    dates = format_dates(dates)
    time = format_time(time)
    group_name = format_group_name(group_name)

    for date in dates:
        db_funcs_for_subjects_db.save_subj(db_name, date, time, group_name, subject)

def parse_work_sheet(work_sheet, db_name):
    time_column = const_time_column
    dates_column = const_dates_column

    groups_column = return_columns_numbers_of_all_groups_cells(work_sheet)
    first_row = find_row_of_first_lesson(work_sheet)
    groups_row = find_number_of_groups_cell_row(work_sheet)

    times = ['9:45', '11:30', '13:30', '15:15', '17:00', '18:40']
    print('ws start')
    for column in groups_column:
        for row in range(first_row, const_quantity_of_rows):
            subject = work_sheet.cell(row = row, column = column).value
            if is_merged(work_sheet, row, column):
                subject = get_value_of_merged_call(work_sheet, row, column)
            if subject == None:
                subject = 'нет предмета'
            time_cell = work_sheet.cell(row = row, column = time_column).value
            if time_cell == None:
                continue
            time = format_time(work_sheet.cell(row = row, column = time_column).value)
            if time in times:
                dates = work_sheet.cell(row = row, column = dates_column).value
                if is_merged(work_sheet, row, dates_column):
                    dates = get_value_of_merged_call(work_sheet, row, dates_column)
                group_name = work_sheet.cell(row = groups_row, column = column).value
                save_subj_in_db(db_name, dates, time, group_name, subject)
            else:
                print(subject)
    print('ws finished')

def return_db_name(file_name):
    if 'zovs_1_kurs' in file_name:
        return 'zovs_1_kurs'
    elif 'zovs_2_kurs' in file_name:
        return 'zovs_2_kurs'
    elif 'zovs_3_kurs' in file_name:
        return 'zovs_3_kurs'
    elif 'zovs_4_kurs' in file_name:
        return 'zovs_4_kurs'
    elif 'lovs_1_kurs' in file_name:
        return 'lovs_1_kurs'
    elif 'lovs_2_kurs' in file_name:
        return 'lovs_2_kurs'
    elif 'lovs_3_kurs' in file_name:
        return 'lovs_3_kurs'
    elif 'lovs_4_kurs' in file_name:
        return 'lovs_4_kurs'

def save_dates_and_times(db_name, dates, times):
    for date in dates:
        for time in times:
            db_funcs_for_subjects_db.save_date_and_time(db_name, date, time)
                            
def get_dates(work_sheet, row, dates_column):
    dates = work_sheet.cell(row = row, column = dates_column).value
    if is_merged(work_sheet, row, dates_column):
        dates = get_value_of_merged_call(work_sheet, row, dates_column)
    dates = format_dates(dates)
    return dates                

def create_dates_and_times_in_db(work_book, db_name):
    dates_column = const_dates_column
    time_column = const_time_column

    for ws in work_book.sheetnames:
        work_sheet = work_book[ws]
        first_row = find_row_of_first_lesson(work_sheet)

        times = []
        for row in range(1, const_quantity_of_rows):
            time_cell = work_sheet.cell(row = row, column = time_column).value
            if time_cell != None and is_time(time_cell):
                time_value = format_time(time_cell)
                if time_value == '9:45':
                    times = []
                    times.append(time_value)
                    dates = get_dates(work_sheet, row, dates_column)
                elif time_value == '11:30' or time_value == '13:30' or time_value == '15:15':
                    times.append(time_value)
                elif time_value == '17:00':
                    times.append(time_value)
                    next_cell = work_sheet.cell(row = row + 1, column = time_column).value
                    next_cell = format_time(next_cell) if is_time(next_cell) else False
                    
                    after_next_cell = work_sheet.cell(row = row + 2, column = time_column).value
                    after_next_cell = format_time(after_next_cell) if is_time(after_next_cell) else False
                        
                    if is_time(next_cell) or is_time(after_next_cell):
                        if next_cell == '9:45' or after_next_cell == '9:45':
                            save_dates_and_times(db_name, dates, times)
                            continue                               
                        elif next_cell == '18:40' or after_next_cell == '18:40':
                            continue
                    elif next_cell == False and after_next_cell == False:
                        save_dates_and_times(db_name, dates, times)
                elif time_value == '18:40':
                    times.append(time_value)
                    save_dates_and_times(db_name, dates, times)

def create_groups_in_db(work_book, db_name):
    work_sheet = work_book[work_book.sheetnames[0]]
    list_of_groups = return_all_groups_names(work_sheet)
    db_funcs_for_subjects_db.save_groups(db_name, list_of_groups)

def parse_work_file(work_file):
    db_name = return_db_name(work_file)
    db_funcs_for_subjects_db.drop_db(db_name)
    db_funcs_for_subjects_db.create_db(db_name)

    work_book = load_workbook(work_file)
    create_groups_in_db(work_book, db_name)
    create_dates_and_times_in_db(work_book, db_name)

    for ws in work_book.sheetnames:
        print(ws)
        work_sheet = work_book[ws]
        parse_work_sheet(work_sheet, db_name)

def parse_work_file_using_name(name):
    work_files = glob.glob('time_tables/full_time_undergraduate/*.xlsx')
    for work_file in work_files:
        if name in work_file:
            db_name = return_db_name(work_file)
            work_book = load_workbook(work_file)
            db_funcs_for_subjects_db.drop_db(db_name)
            create_groups_in_db(work_book, db_name)
            create_dates_and_times_in_db(work_book, db_name)
        
            for ws in work_book.sheetnames:
                print(ws)
                work_sheet = work_book[ws]
                parse_work_sheet(work_sheet, db_name)

def run_parser():
    work_files = glob.glob('time_tables/full_time_undergraduate/*.xlsx')
    for work_file in work_files:
        print(work_file)
        parse_work_file(work_file)
        
if __name__ == "__main__":
    run_parser()



