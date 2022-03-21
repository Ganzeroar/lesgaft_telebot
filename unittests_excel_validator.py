import unittest
from openpyxl import Workbook, load_workbook, utils
import glob

import excel_validator


class Test_excel_validator(unittest.TestCase):
    # потенциально валидировать имя файла + 
    # валидировать количество групп в расписании?
    # валидировать на поиск пустых строк?
    # ограничить количество групп в каждом расписании, тем самым контроллируя строки?
    # провверять правильность день дата время?
    # проверять первую ячейку по константам на None
    # добавить поиск пар в 18:40
    @unittest.skip("passed")
    def test_take_one_correct_and_one_incorrect_sheetname_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        result = obj.run_validator('test_time_tables/full_time_undergraduate')
        self.assertEqual('Имя файла ОК\nИмена листов ОК\nЛист 21.03. - 26.03.\nСтруктура групп ОК\nСруктура дат ОК\nСруктура дней ОК\nСруктура времени ОК\nЛист 21.03. - 26.03.\nИмена групп ОК\nДаты ОК\nИмена дней ОК\nВремя ОК\n', result)

    #def test_take_full_file_return_correct_message(self):
    #    obj = excel_validator.Excel_validator()
    #    result = obj.run_validator('test_time_tables/full_time_undergraduate/fulltest')
    #    self.assertEqual('ОК\n', result)

class Test_check_worksheet_name(unittest.TestCase):
    def test_take_correct_sheetnames_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        result = obj.check_worksheet_names(['24.01. - 29.01.', '31.01. - 05.02.'])
        self.assertEqual('Имена листов ОК\n', result)

    def test_take_one_incorrect_sheetname_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        result = obj.check_worksheet_names(['24.01 - 29.01', '31.01. - 05.02.'])
        self.assertEqual('Ошибка в имени листа 24.01 - 29.01\n', result)

    def test_take_all_incorrect_sheetname_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        result = obj.check_worksheet_names(['24.01 - 29.01', '31.01 - 05.02'])
        self.assertEqual('Ошибка в имени листа 24.01 - 29.01\nОшибка в имени листа 31.01 - 05.02\n', result)

class Test_check_file_name(unittest.TestCase):
    def test_take_correct_filename_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        result = obj.check_file_name('2_LOVS_93')
        self.assertEqual('Имя файла ОК\n', result)

    def test_take_incorrect_filename_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        result = obj.check_file_name('2_LOV_93')
        self.assertEqual('Ошибка в имени файла 2_LOV_93\n', result)

class Test_check_group_struct(unittest.TestCase):
    def test_take_correct_group_struct_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
        work_book = load_workbook(file_name)
        worksheet = work_book[work_book.sheetnames[0]]
        result = obj.check_group_struct(worksheet, '2_LOVS')
        self.assertEqual('Структура групп ОК\n', result)

    @unittest.skip("passed")
    def test_take_incorrect_group_struct_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
        work_book = load_workbook(file_name)
        worksheet = work_book[work_book.sheetnames[0]]
        result = obj.check_group_struct(worksheet, '2_LOVS')
        self.assertEqual('Ошибка в D4\n', result)

class Test_check_date_struct(unittest.TestCase):
    def test_take_correct_date_struct_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
        work_book = load_workbook(file_name)
        worksheet = work_book[work_book.sheetnames[0]]
        result = obj.check_date_struct(worksheet, '2_LOVS')
        self.assertEqual('Сруктура дат ОК\n', result)

class Test_check_day_struct(unittest.TestCase):
    def test_take_correct_day_struct_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
        work_book = load_workbook(file_name)
        worksheet = work_book[work_book.sheetnames[0]]
        result = obj.check_day_struct(worksheet, '2_LOVS')
        self.assertEqual('Сруктура дней ОК\n', result)

class Test_check_time_struct(unittest.TestCase):
    def test_take_correct_time_struct_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
        work_book = load_workbook(file_name)
        worksheet = work_book[work_book.sheetnames[0]]
        result = obj.check_time_struct(worksheet, '2_LOVS')
        self.assertEqual('Сруктура времени ОК\n', result)

class Test_check_group_numbers(unittest.TestCase):
    def test_take_correct_group_names_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
        work_book = load_workbook(file_name)
        worksheet = work_book[work_book.sheetnames[0]]
        result = obj.check_group_numbers(worksheet, '2_LOVS')
        self.assertEqual('Имена групп ОК\n', result)

class Test_check_date_column(unittest.TestCase):
    def test_take_correct_date_column_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
        work_book = load_workbook(file_name)
        worksheet = work_book[work_book.sheetnames[0]]
        result = obj.check_date_column(worksheet, '2_LOVS')
        self.assertEqual('Даты ОК\n', result)

class Test_check_day_of_week_column(unittest.TestCase):
    def test_take_correct_day_column_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
        work_book = load_workbook(file_name)
        worksheet = work_book[work_book.sheetnames[0]]
        result = obj.check_day_of_week_column(worksheet, '2_LOVS')
        self.assertEqual('Имена дней ОК\n', result)

class Test_check_time_column(unittest.TestCase):
    def test_take_correct_time_column_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
        work_book = load_workbook(file_name)
        worksheet = work_book[work_book.sheetnames[0]]
        result = obj.check_time_column(worksheet, '2_LOVS')
        self.assertEqual('Время ОК\n', result)


#class Test_check_constants_of_excel_spreadsheet(unittest.TestCase):
#    def test_take_correct_day_struct_return_correct_message(self):
#        obj = excel_validator.Excel_validator()
#        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
#        work_book = load_workbook(file_name)
#        worksheet = work_book[work_book.sheetnames[0]]
#        result = obj.check_constants_of_excel_spreadsheet(worksheet, '2_LOVS')
#        self.assertEqual('Структура групп ОК\nСруктура дат ОК\nСруктура дней ОК\nСруктура времени ОК\n', result)

class Test_check_structure(unittest.TestCase):
    def test_take_correct_struct_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
        work_book = load_workbook(file_name)
        result = obj.check_structure(work_book, '2_LOVS')
        self.assertEqual('Лист 21.03. - 26.03.\nСтруктура групп ОК\nСруктура дат ОК\nСруктура дней ОК\nСруктура времени ОК\n', result)

class Test_check_content_of_servise_cells(unittest.TestCase):
    def test_take_correct_content_of_servise_cells_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
        work_book = load_workbook(file_name)
        result = obj.check_content_of_servise_cells(work_book, '2_LOVS')
        self.assertEqual('Лист 21.03. - 26.03.\nИмена групп ОК\nДаты ОК\nИмена дней ОК\nВремя ОК\n', result)

@unittest.skip("не решили что делать с ячейкой ловс 2 м24")
class Test_check_class_schedule(unittest.TestCase):
    def test_take_correct_content_cells_return_correct_message(self):
        obj = excel_validator.Excel_validator()
        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/*.xlsx')[0]
        work_book = load_workbook(file_name)
        result = obj.check_class_schedule(work_book, '2_LOVS')
        self.assertEqual('Лист 21.03. - 26.03.\nПредметы ОК\n', result)

#class Test_check_class_schedule_full(unittest.TestCase):
#    def test_take_correct_content_cells_return_correct_message(self):
#        obj = excel_validator.Excel_validator()
#        file_name = glob.glob('time_tables/test_time_tables/full_time_undergraduate/fulltest/*.xlsx')[0]
#        work_book = load_workbook(file_name)
#        result = obj.check_class_schedule(work_book, 'lovs_1')
#        self.assertEqual('Лист 21.03. - 26.03.\nПредметы ОК\n', result)


#class Test_check_group_name(unittest.TestCase):
#    def test_take_correct_group_names_return_correct_message(self):
#        obj = excel_validator.Excel_validator()
#        work_files = glob.glob(f'time_tables/test_time_tables/full_time_undergraduate/*.xlsx')
#        work_book = load_workbook(work_files[0])
#        result = obj.check_group_numbers(work_book)
#        self.assertEqual('Имена групп ОК', result)

        
if __name__ == '__main__':
    unittest.main()